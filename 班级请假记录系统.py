"""
班级请假记录系统
功能包括：学生名单管理、请假录入、数据统计、Excel导出等
"""

import os
import sys
import json
import datetime
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import importlib
import tkintercalendar
importlib.reload(tkintercalendar)
CalendarWidget = tkintercalendar.Calendar
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import shutil

# 获取程序运行目录
if getattr(sys, 'frozen', False):
    # 打包后的以
    """动画效果辅助类"""
    
    @staticmethod
    def fade_in(widget, duration=300, callback=None):
        """淡入效果"""
        steps = 20
        delay = duration // steps
        
        def step(current_step):
            if current_step <= steps:
                alpha = current_step / steps
                widget.attributes('-alpha', alpha)
                widget.after(delay, lambda: step(current_step + 1))
            else:
                widget.attributes('-alpha', 1.0)
                if callback:
                    callback()
        
        widget.attributes('-alpha', 0.0)
        step(0)
    
    @staticmethod
    def slide_in(widget, direction='left', duration=300, callback=None):
        """滑入效果"""
        steps = 20
        delay = duration // steps
        
        # 获取窗口位置
        x = widget.winfo_x()
        y = widget.winfo_y()
        width = widget.winfo_width()
        height = widget.winfo_height()
        
        # 设置初始位置
        if direction == 'left':
            start_x = x - width
            start_y = y
        elif direction == 'right':
            start_x = x + width
            start_y = y
        elif direction == 'top':
            start_x = x
            start_y = y - height
        elif direction == 'bottom':
            start_x = x
            start_y = y + height
        else:
            start_x = x
            start_y = y
        
        widget.geometry(f"+{start_x}+{start_y}")
        
        def step(current_step):
            if current_step <= steps:
                progress = current_step / steps
                # 使用缓动函数
                ease = progress * (2 - progress)
                current_x = int(start_x + (x - start_x) * ease)
                current_y = int(start_y + (y - start_y) * ease)
                widget.geometry(f"+{current_x}+{current_y}")
                widget.after(delay, lambda: step(current_step + 1))
            else:
                widget.geometry(f"+{x}+{y}")
                if callback:
                    callback()
        
        step(0)
    
    @staticmethod
    def pulse(widget, color_bg, color_fg, duration=1000, repeat=3):
        """脉冲效果"""
        half_duration = duration // 2
        steps = 10
        delay = half_duration // steps
        
        original_bg = widget.cget('background')
        original_fg = widget.cget('foreground')
        
        def pulse_step(current_step, forward):
            if current_step <= steps:
                progress = current_step / steps
                if forward:
                    widget.configure(background=color_bg, foreground=color_fg)
                else:
                    widget.configure(background=original_bg, foreground=original_fg)
                widget.after(delay, lambda: pulse_step(current_step + 1, forward))
            else:
                if forward:
                    widget.after(half_duration, lambda: pulse_step(0, False))
                else:
                    widget.configure(background=original_bg, foreground=original_fg)
        
        pulse_step(0, True)
    
    @staticmethod
    def highlight_border(widget, color='#3498DB', duration=1000):
        """边框高亮效果"""
        widget.configure(relief='solid', borderwidth=2, highlightbackground=color, highlightthickness=2)
        
        def restore():
            widget.configure(relief='flat', borderwidth=0, highlightthickness=0)
        
        widget.after(duration, restore)


class StudentManager:
    """学生名单管理"""

    def __init__(self, data_file: str = "students.json"):
        # 确保data文件夹存在
        data_dir = 'data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        self.data_file = os.path.join(data_dir, data_file)
        self.students = []
        self.load_students()
    
    def load_students(self):
        """加载学生名单"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    self.students = json.load(f)
            except:
                self.students = []
        else:
            # 首次运行，初始化空名单
            self.students = []
            self.save_students()
    
    def save_students(self):
        """保存学生名单"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.students, f, ensure_ascii=False, indent=2)
    
    def add_student(self, name: str) -> bool:
        """添加学生"""
        if not name or name in self.students:
            return False
        self.students.append(name)
        self.students.sort()  # 按拼音排序
        self.save_students()
        return True
    
    def remove_student(self, name: str) -> bool:
        """删除学生"""
        if name in self.students:
            self.students.remove(name)
            self.save_students()
            return True
        return False
    
    def batch_import(self, names: List[str]) -> int:
        """批量导入学生"""
        count = 0
        for name in names:
            if name and name not in self.students:
                self.students.append(name)
                count += 1
        self.students.sort()
        self.save_students()
        return count
    
    def get_students(self) -> List[str]:
        """获取学生列表（按拼音排序）"""
        from pypinyin import lazy_pinyin

        return sorted(self.students, key=lambda x: ''.join(lazy_pinyin(x)))


class LeaveRecordManager:
    """请假记录管理（改进版 - 添加原子性保护和线程安全）"""

    def __init__(self, data_file: str = "leave_records.json"):
        # 确保data文件夹存在
        data_dir = 'data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        self.data_file = os.path.join(data_dir, data_file)
        self.records = {}  # {date: {name: {"type": "half"/"full"}}}
        self.load_records()

        # 添加数据锁，防止并发写入
        self._lock = threading.Lock()

    def load_records(self):
        """加载请假记录"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    self.records = json.load(f)
            except:
                self.records = {}
        else:
            self.records = {}

    def save_records(self):
        """保存请假记录（改进版 - 添加原子性保护）"""
        with self._lock:
            # 创建临时文件
            temp_file = self.data_file + '.tmp'

            try:
                # 写入临时文件
                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(self.records, f, ensure_ascii=False, indent=2)

                # 使用原子操作替换原文件
                if os.path.exists(self.data_file):
                    os.replace(temp_file, self.data_file)
                else:
                    os.rename(temp_file, self.data_file)

            except Exception as e:
                # 清理临时文件
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                raise e

    def add_leave(self, date: str, name: str, leave_type: str):
        """添加请假记录（改进版 - 不立即保存）"""
        with self._lock:
            if date not in self.records:
                self.records[date] = {}
            self.records[date][name] = {"type": leave_type}
            # 移除立即保存，由调用方统一保存

    def remove_leave(self, date: str, name: str):
        """删除请假记录（改进版 - 不立即保存）"""
        with self._lock:
            if date in self.records and name in self.records[date]:
                del self.records[date][name]
                if not self.records[date]:
                    del self.records[date]
                # 移除立即保存，由调用方统一保存
    
    def update_leave(self, date: str, name: str, leave_type: str):
        """更新请假记录"""
        if date in self.records and name in self.records[date]:
            self.records[date][name]["type"] = leave_type
            self.save_records()
    
    def get_leave_records(self, date: str) -> Dict[str, str]:
        """获取某天的请假记录"""
        return self.records.get(date, {})
    
    def get_all_dates(self) -> List[str]:
        """获取所有有记录的日期"""
        return sorted(self.records.keys())
    
    def get_frequent_leavers(self, days: int = 5, threshold: int = 3) -> List[str]:
        """获取常请假的学生"""
        end_date = datetime.datetime.now()
        start_date = end_date - datetime.timedelta(days=days)
        
        leave_counts = defaultdict(int)
        
        for date_str, records in self.records.items():
            try:
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                if start_date <= date <= end_date:
                    for name in records:
                        leave_counts[name] += 1
            except:
                continue
        
        return [name for name, count in leave_counts.items() if count >= threshold]
    
    def get_student_leave_history(self, name: str) -> Dict[str, str]:
        """获取某学生的请假历史"""
        history = {}
        for date_str, records in self.records.items():
            if name in records:
                history[date_str] = records[name]["type"]
        return sorted(history.items())
    
    def get_statistics(self, start_date: str, end_date: str) -> Dict:
        """获取统计数据"""
        stats = {
            "total_days": 0,
            "total_half_days": 0,
            "total_full_days": 0,
            "weekdays": {"half_days": 0, "full_days": 0, "students": []},
            "saturdays": {"half_days": 0, "full_days": 0, "students": []},
            "sundays": {"half_days": 0, "full_days": 0, "students": []},
            "daily": {}
        }
        
        for date_str, records in self.records.items():
            if start_date <= date_str <= end_date:
                try:
                    date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                    weekday = date.weekday()  # 0=周一, 6=周日
                    
                    half_count = 0
                    full_count = 0
                    students = []
                    
                    for name, record in records.items():
                        if record["type"] == "half":
                            half_count += 1
                            stats["total_half_days"] += 1
                        else:
                            full_count += 1
                            stats["total_full_days"] += 1
                        students.append(name)
                    
                    stats["total_days"] += 1
                    stats["daily"][date_str] = {
                        "half_days": half_count,
                        "full_days": full_count,
                        "students": students
                    }
                    
                    if weekday == 6:  # 周日
                        stats["sundays"]["half_days"] += half_count
                        stats["sundays"]["full_days"] += full_count
                        stats["sundays"]["students"].extend(students)
                    elif weekday == 5:  # 周六
                        stats["saturdays"]["half_days"] += half_count
                        stats["saturdays"]["full_days"] += full_count
                        stats["saturdays"]["students"].extend(students)
                    else:  # 工作日
                        stats["weekdays"]["half_days"] += half_count
                        stats["weekdays"]["full_days"] += full_count
                        stats["weekdays"]["students"].extend(students)
                except:
                    continue
        
        # 去重学生名单
        stats["weekdays"]["students"] = list(set(stats["weekdays"]["students"]))
        stats["saturdays"]["students"] = list(set(stats["saturdays"]["students"]))
        stats["sundays"]["students"] = list(set(stats["sundays"]["students"]))
        
        return stats
    
    def get_student_statistics(self, name: str, start_date: str, end_date: str) -> Dict:
        """获取某学生的请假统计"""
        stats = {
            "total_half_days": 0,
            "total_full_days": 0,
            "weekdays": {"half_days": 0, "full_days": 0, "dates": []},
            "saturdays": {"half_days": 0, "full_days": 0, "dates": []},
            "sundays": {"half_days": 0, "full_days": 0, "dates": []},
            "records": []
        }
        
        for date_str, records in self.records.items():
            if start_date <= date_str <= end_date and name in records:
                try:
                    date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                    weekday = date.weekday()
                    leave_type = records[name]["type"]
                    
                    record_info = {
                        "date": date_str,
                        "type": leave_type,
                        "weekday": weekday
                    }
                    stats["records"].append(record_info)
                    
                    if leave_type == "half":
                        stats["total_half_days"] += 1
                    else:
                        stats["total_full_days"] += 1
                    
                    if weekday == 6:
                        stats["sundays"]["half_days"] += 1 if leave_type == "half" else 0
                        stats["sundays"]["full_days"] += 1 if leave_type == "full" else 0
                        stats["sundays"]["dates"].append(date_str)
                    elif weekday == 5:
                        stats["saturdays"]["half_days"] += 1 if leave_type == "half" else 0
                        stats["saturdays"]["full_days"] += 1 if leave_type == "full" else 0
                        stats["saturdays"]["dates"].append(date_str)
                    else:
                        stats["weekdays"]["half_days"] += 1 if leave_type == "half" else 0
                        stats["weekdays"]["full_days"] += 1 if leave_type == "full" else 0
                        stats["weekdays"]["dates"].append(date_str)
                except:
                    continue
        
        return stats


class CalendarWidget:
    """日历组件"""
    
    def __init__(self, parent, on_date_select=None):
        self.parent = parent
        self.on_date_select = on_date_select
        self.selected_date = None
        self.highlighted_dates = set()
        self.current_year = datetime.datetime.now().year
        self.current_month = datetime.datetime.now().month
        
        self.create_widgets()
    
    def create_widgets(self):
        """创建日历组件"""
        # 月份导航
        nav_frame = tk.Frame(self.parent)
        nav_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.prev_btn = tk.Button(nav_frame, text="<", width=3, command=self.prev_month)
        self.prev_btn.pack(side=tk.LEFT)
        
        self.month_label = tk.Label(nav_frame, text="", font=("Arial", 12, "bold"))
        self.month_label.pack(side=tk.LEFT, expand=True)
        
        self.next_btn = tk.Button(nav_frame, text=">", width=3, command=self.next_month)
        self.next_btn.pack(side=tk.RIGHT)
        
        # 星期标题
        week_frame = tk.Frame(self.parent)
        week_frame.pack(fill=tk.X)
        
        weekdays = ["日", "一", "二", "三", "四", "五", "六"]
        for i, day in enumerate(weekdays):
            label = tk.Label(week_frame, text=day, width=5, font=("Arial", 10, "bold"))
            label.grid(row=0, column=i, padx=1, pady=1)
        
        # 日历主体
        self.calendar_frame = tk.Frame(self.parent)
        self.calendar_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.day_buttons = {}
        self.update_calendar()
    
    def update_calendar(self):
        """更新日历显示"""
        # 清空现有按钮
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()
        
        # 更新月份标签
        month_names = ["一月", "二月", "三月", "四月", "五月", "六月",
                      "七月", "八月", "九月", "十月", "十一月", "十二月"]
        self.month_label.config(text=f"{self.current_year}年 {month_names[self.current_month-1]}")
        
        # 获取该月第一天是星期几
        first_day = datetime.datetime(self.current_year, self.current_month, 1)
        start_weekday = first_day.weekday() + 1  # 0=周一, 6=周日, 转换为0=周日, 6=周六
        
        # 获取该月总天数
        if self.current_month == 12:
            next_month = datetime.datetime(self.current_year + 1, 1, 1)
        else:
            next_month = datetime.datetime(self.current_year, self.current_month + 1, 1)
        total_days = (next_month - first_day).days
        
        # 创建日历按钮
        day = 1
        for row in range(6):
            for col in range(7):
                if row == 0 and col < start_weekday:
                    continue
                if day > total_days:
                    break
                
                date_str = f"{self.current_year}-{self.current_month:02d}-{day:02d}"
                
                # 检查是否是高亮日期
                bg_color = "white"
                fg_color = "black"
                if date_str in self.highlighted_dates:
                    bg_color = "#FFD700"  # 金色
                    fg_color = "black"
                
                # 检查是否是选中日期
                if date_str == self.selected_date:
                    bg_color = "#4CAF50"  # 绿色
                    fg_color = "white"
                
                btn = tk.Button(self.calendar_frame, text=str(day), width=5, height=2,
                               bg=bg_color, fg=fg_color,
                               command=lambda d=date_str: self.select_date(d))
                btn.grid(row=row, column=col, padx=1, pady=1)
                
                day += 1
    
    def prev_month(self):
        """上个月"""
        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        self.update_calendar()
    
    def next_month(self):
        """下个月"""
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        self.update_calendar()
    
    def select_date(self, date_str: str):
        """选择日期"""
        self.selected_date = date_str
        self.update_calendar()
        if self.on_date_select:
            self.on_date_select(date_str)
    
    def highlight_dates(self, dates: List[str]):
        """高亮显示日期"""
        self.highlighted_dates = set(dates)
        self.update_calendar()
    
    def set_selected_date(self, date_str: str):
        """设置选中日期"""
        self.selected_date = date_str
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            self.current_year = date.year
            self.current_month = date.month
        except:
            pass
        self.update_calendar()
    
    def get_selected_date(self) -> str:
        """获取选中日期"""
        return self.selected_date


class LeaveRecordApp:
    """请假记录应用主类"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("班级请假记录系统 v1.0.1")
        self.root.state('zoomed')  # 最大化窗口
        self.root.minsize(1200, 800)

        # 设置窗口图标
        try:
            icon_path = os.path.join(BASE_DIR, 'calendar_icon.ico')
            if os.path.exists(icon_path):
                # 尝试使用iconbitmap方法
                try:
                    self.root.iconbitmap(icon_path)
                except:
                    # 如果失败,尝试使用PIL加载PNG图标并转换为tkinter格式
                    try:
                        from PIL import Image, ImageTk
                        png_icon_path = os.path.join(BASE_DIR, 'calendar_icon.png')
                        if os.path.exists(png_icon_path):
                            # 加载PNG图标
                            img = Image.open(png_icon_path)
                            # 调整大小为32x32(适合窗口图标)
                            img = img.resize((32, 32), Image.Resampling.LANCZOS)
                            # 转换为tkinter格式
                            icon_image = ImageTk.PhotoImage(img)
                            self.root.iconphoto(False, icon_image)
                    except:
                        pass
        except:
            pass  # 如果图标文件不存在,忽略错误

        # 设置窗口背景
        try:
            self.root.configure(bg=self.colors['bg'])
        except:
            pass

        # 设置样式
        self.setup_styles()

        # 初始化管理器
        self.student_manager = StudentManager()
        self.leave_manager = LeaveRecordManager()

        # 初始化学生请假类型字典
        self.student_leave_types = {}  # {name: "full" or "half" or None}

        # 标记是否有未保存的修改
        self.has_unsaved_changes = False

        # 标记是否正在保存
        self.is_saving = False

        # 日历更新防抖定时器
        self._calendar_update_timer = None

        # 添加关闭窗口事件处理
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # 创建UI
        self.create_ui()

        # 延迟加载初始数据，优化启动速度
        self.root.after(100, self.load_initial_data)

    def on_closing(self):
        """关闭窗口时的处理"""
        # 保存设置
        self.save_settings()

        if self.has_unsaved_changes:
            if messagebox.askyesno("未保存的修改", "检测到有未保存的请假记录，是否保存？"):
                self.save_leave_record()
        self.root.destroy()
    
    def setup_styles(self):
        """设置样式"""
        style = ttk.Style()
        
        # 设置主题
        style.theme_use('clam')
        
        # 设置颜色 - 现代配色方案
        colors = {
            'bg': '#F5F7FA',           # 浅灰蓝背景
            'fg': '#2C3E50',           # 深灰蓝文字
            'accent': '#3498DB',       # 主题蓝
            'accent_hover': '#2980B9', # 主题蓝悬停
            'success': '#27AE60',      # 成功绿
            'warning': '#F39C12',      # 警告橙
            'danger': '#E74C3C',       # 危险红
            'white': '#FFFFFF',        # 白色
            'light_gray': '#ECF0F1',   # 浅灰
            'border': '#BDC3C7'        # 边框色
        }
        
        # 配置通用样式
        style.configure('TFrame', background=colors['bg'])
        style.configure('TLabel', background=colors['bg'], foreground=colors['fg'], font=('Microsoft YaHei', 10))
        style.configure('Header.TLabel', background=colors['accent'], foreground=colors['white'], 
                       font=('Microsoft YaHei', 12, 'bold'))
        
        # 配置按钮样式
        style.configure('TButton', 
                       font=('Microsoft YaHei', 10),
                       padding=8,
                       background=colors['accent'],
                       foreground=colors['white'],
                       borderwidth=0)
        style.map('TButton',
                 background=[('active', colors['accent_hover']),
                           ('pressed', colors['accent_hover'])])
        
        # 配置按钮变体
        style.configure('Success.TButton', background=colors['success'])
        style.map('Success.TButton',
                 background=[('active', '#229954'),
                           ('pressed', '#229954')])
        
        style.configure('Warning.TButton', background=colors['warning'])
        style.map('Warning.TButton',
                 background=[('active', '#D68910'),
                           ('pressed', '#D68910')])
        
        style.configure('Danger.TButton', background=colors['danger'])
        style.map('Danger.TButton',
                 background=[('active', '#CB4335'),
                           ('pressed', '#CB4335')])
        
        # 配置Treeview样式
        style.configure('Treeview', 
                       rowheight=28,
                       font=('Microsoft YaHei', 10),
                       background=colors['white'],
                       fieldbackground=colors['white'],
                       borderwidth=0)
        style.configure('Treeview.Heading', 
                       font=('Microsoft YaHei', 10, 'bold'),
                       background=colors['light_gray'],
                       foreground=colors['fg'],
                       relief='flat')
        style.map('Treeview',
                 background=[('selected', colors['accent'])],
                 foreground=[('selected', colors['white'])])
        
        # 配置Entry样式
        style.configure('TEntry',
                       fieldbackground=colors['white'],
                       borderwidth=1,
                       relief='solid')
        style.map('TEntry',
                 bordercolor=[('focus', colors['accent'])])
        
        # 配置Combobox样式
        style.configure('TCombobox',
                       fieldbackground=colors['white'],
                       borderwidth=1,
                       relief='solid')
        style.map('TCombobox',
                 bordercolor=[('focus', colors['accent'])])
        
        # 配置Notebook样式
        style.configure('TNotebook',
                       background=colors['bg'],
                       borderwidth=0)
        style.configure('TNotebook.Tab',
                       background=colors['light_gray'],
                       foreground=colors['fg'],
                       padding=[20, 8],
                       font=('Microsoft YaHei', 10))
        style.map('TNotebook.Tab',
                 background=[('selected', colors['accent']),
                           ('active', colors['light_gray'])],
                 foreground=[('selected', colors['white'])])
        
        # 保存颜色供后续使用
        self.colors = colors

    def create_ui(self):
        """创建UI界面 - 全新宽松设计"""
        # 创建主容器
        main_container = ttk.Frame(self.root, style='TFrame')
        main_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # 顶部工具栏
        toolbar = tk.Frame(main_container, bg=self.colors['light_gray'], height=60)
        toolbar.pack(fill=tk.X, pady=(0, 15))
        toolbar.pack_propagate(False)
        self.create_toolbar(toolbar)

        # 创建左右两栏布局
        content_frame = tk.Frame(main_container, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧面板（日历和操作）
        left_panel = tk.Frame(content_frame, bg=self.colors['white'], width=360)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 15), pady=0)
        left_panel.pack_propagate(False)

        # 右侧面板（学生列表和统计）
        right_panel = tk.Frame(content_frame, bg=self.colors['white'])
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=0, pady=0)

        # 创建左侧面板内容
        self.create_left_panel(left_panel)

        # 创建右侧面板内容
        self.create_right_panel(right_panel)

        # 创建底部状态栏
        self.status_bar = tk.Label(self.root, text="就绪",
                                  bg=self.colors['light_gray'], fg=self.colors['fg'],
                                  font=('Microsoft YaHei', 11),
                                  anchor='w', padx=10, pady=5)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def update_status(self, message):
        """更新状态栏信息"""
        if hasattr(self, 'status_bar'):
            self.status_bar.config(text=message)

    def create_toolbar(self, parent):
        """创建顶部工具栏（宽松设计 - 添加学生管理）"""
        # 标题
        title_label = tk.Label(parent, text="班级请假记录系统", 
                             font=('Microsoft YaHei', 16, 'bold'),
                             bg=self.colors['light_gray'], fg=self.colors['fg'])
        title_label.pack(side=tk.LEFT, padx=15)
        
        # 分隔线
        separator = tk.Frame(parent, bg=self.colors['accent'], width=2)
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=8)
        
        # 当前日期显示
        self.date_var = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        date_label = tk.Label(parent, text=f"📅 日期: {self.date_var.get()}", 
                            font=('Microsoft YaHei', 11),
                            bg=self.colors['light_gray'], fg=self.colors['fg'])
        date_label.pack(side=tk.LEFT, padx=15)
        
        # 学生管理按钮
        add_btn = tk.Button(parent, text="➕ 添加学生",
                          command=self.show_add_student_dialog,
                          bg=self.colors['accent'], fg=self.colors['white'],
                          font=('Segoe UI Symbol', 10, 'bold'), relief='flat',
                          padx=12, pady=6, cursor='hand2', bd=0)
        add_btn.pack(side=tk.LEFT, padx=(0, 8))
        self._add_button_hover_effect(add_btn, self.colors['accent'], self.colors['accent_hover'])

        import_btn = tk.Button(parent, text="📥 导入学生",
                             command=self.show_batch_import_dialog,
                             bg=self.colors['warning'], fg=self.colors['white'],
                             font=('Segoe UI Symbol', 10, 'bold'), relief='flat',
                             padx=12, pady=6, cursor='hand2', bd=0)
        import_btn.pack(side=tk.LEFT, padx=(0, 8))
        self._add_button_hover_effect(import_btn, self.colors['warning'], '#D68910')

        remove_btn = tk.Button(parent, text="❌ 删除学生",
                             command=self.show_remove_student_dialog,
                             bg=self.colors['danger'], fg=self.colors['white'],
                             font=('Segoe UI Symbol', 10, 'bold'), relief='flat',
                             padx=12, pady=6, cursor='hand2', bd=0)
        remove_btn.pack(side=tk.LEFT, padx=(15, 8))
        self._add_button_hover_effect(remove_btn, self.colors['danger'], '#CB4335')
        
        # 操作按钮 - 修复emoji和文字对齐问题，使用flat样式避免闪烁
        self.save_btn = tk.Button(parent, text="💾 保存",
                           command=self.save_leave_record,
                           bg=self.colors['success'], fg=self.colors['white'],
                           font=('Segoe UI Emoji', 11, 'bold'), relief='flat',
                           padx=20, pady=8, cursor='hand2', bd=0,
                           compound='left', anchor='center')
        self.save_btn.pack(side=tk.RIGHT, padx=(0, 12))
        self._add_button_hover_effect(self.save_btn, self.colors['success'], '#229954')

        clear_btn = tk.Button(parent, text="🔄 清空",
                            command=self.clear_selection,
                            bg=self.colors['warning'], fg=self.colors['white'],
                            font=('Segoe UI Emoji', 11, 'bold'), relief='flat',
                            padx=20, pady=8, cursor='hand2', bd=0,
                            compound='left', anchor='center')
        clear_btn.pack(side=tk.RIGHT, padx=(0, 8))
        self._add_button_hover_effect(clear_btn, self.colors['warning'], '#D68910')
    
    def _add_button_hover_effect(self, button, normal_color, hover_color):
        """为按钮添加悬停效果"""
        def on_enter(event):
            button.configure(bg=hover_color)
        
        def on_leave(event):
            button.configure(bg=normal_color)
        
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)
    
    def _animate_startup(self):
        """启动动画 - 优化版本"""
        # 窗口淡入效果
        self.root.attributes('-alpha', 0.0)

        def fade_in(step, max_steps=25):
            if step <= max_steps:
                # 使用缓动函数
                progress = step / max_steps
                ease = progress * (2 - progress)  # ease-out
                alpha = ease
                self.root.attributes('-alpha', alpha)
                self.root.after(20, lambda: fade_in(step + 1))
            else:
                self.root.attributes('-alpha', 1.0)

        fade_in(0)

    def _animate_success(self, message="操作成功！"):
        """成功动画效果 - 1秒版本"""
        # 创建一个临时的成功提示窗口
        success_window = tk.Toplevel(self.root)
        success_window.title("")
        success_window.geometry("320x90")
        success_window.overrideredirect(True)
        success_window.attributes('-topmost', True)
        success_window.attributes('-alpha', 0.0)

        # 居中显示
        x = self.root.winfo_x() + (self.root.winfo_width() - 320) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 90) // 2
        success_window.geometry(f"+{x}+{y}")

        # 创建内容
        frame = tk.Frame(success_window, bg=self.colors['success'], padx=25, pady=25)
        frame.pack(fill=tk.BOTH, expand=True)

        label = tk.Label(frame, text=f"✓ {message}",
                        font=('Segoe UI Emoji', 13, 'bold'),
                        bg=self.colors['success'], fg=self.colors['white'])
        label.pack()

        # 动画效果 - 1秒版本（约67帧，每帧15ms）
        def show_and_hide(step, max_steps=67):
            if step <= max_steps:
                if step <= 20:
                    # 淡入和放大 - 前20帧（约300ms）
                    progress = step / 20
                    alpha = progress
                    scale = 0.8 + progress * 0.2
                    success_window.attributes('-alpha', alpha)
                    # 缩放窗口
                    current_width = int(320 * scale)
                    current_height = int(90 * scale)
                    x = self.root.winfo_x() + (self.root.winfo_width() - current_width) // 2
                    y = self.root.winfo_y() + (self.root.winfo_height() - current_height) // 2
                    success_window.geometry(f'{current_width}x{current_height}+{x}+{y}')
                elif step >= 47:
                    # 淡出 - 后20帧（约300ms）
                    progress = (step - 47) / 20
                    alpha = 1 - progress
                    success_window.attributes('-alpha', alpha)

                success_window.after(15, lambda: show_and_hide(step + 1))
            else:
                success_window.destroy()

        show_and_hide(0)
    
    def _animate_pulse(self, widget, color1, color2):
        """脉冲动画效果"""
        steps = 3
        
        def pulse(step, forward):
            if step <= steps:
                if forward:
                    widget.config(bg=color1)
                else:
                    widget.config(bg=color2)
                self.root.after(150, lambda: pulse(step + 1, not forward))
            else:
                if forward:
                    widget.config(bg=color1)
        
        pulse(0, True)
    
    def _animate_bounce(self, widget, times=3):
        """弹跳动画效果"""
        for i in range(times):
            self.root.after(i * 200, lambda: widget.config(relief='raised', borderwidth=3))
            self.root.after(i * 200 + 100, lambda: widget.config(relief='flat', borderwidth=0))
    
    def _animate_shake(self, widget, times=2):
        """抖动动画效果"""
        original_x = widget.winfo_x()
        
        def shake(step):
            if step < times * 4:
                offset = 5 if step % 2 == 0 else -5
                widget.place(x=original_x + offset)
                widget.after(50, lambda: shake(step + 1))
            else:
                widget.place(x=original_x)
        
        shake(0)
    
    def _animate_glow(self, widget, color='#3498DB'):
        """发光动画效果"""
        original_bg = widget.cget('background')
        original_fg = widget.cget('foreground')
        
        def glow(step, max_steps=10, increasing=True):
            if increasing:
                if step < max_steps:
                    widget.config(background=color, foreground='white')
                    widget.after(30, lambda: glow(step + 1, max_steps, True))
                else:
                    widget.after(100, lambda: glow(0, max_steps, False))
            else:
                if step < max_steps:
                    widget.config(background=original_bg, foreground=original_fg)
                    widget.after(30, lambda: glow(step + 1, max_steps, False))
                else:
                    widget.config(background=original_bg, foreground=original_fg)
        
        glow(0, 8, True)
    
    def _animate_slide_in(self, widget, direction='left'):
        """滑入动画效果"""
        widget.place_forget()
        widget.update()
        
        x = widget.winfo_x()
        y = widget.winfo_y()
        width = widget.winfo_width()
        height = widget.winfo_height()
        
        # 设置初始位置
        if direction == 'left':
            start_x = x - width
        elif direction == 'right':
            start_x = x + width
        elif direction == 'top':
            start_x = x
            y = y - height
        elif direction == 'bottom':
            start_x = x
            y = y + height
        else:
            start_x = x
        
        widget.place(x=start_x, y=y)
        
        def slide(step, max_steps=15):
            if step <= max_steps:
                progress = step / max_steps
                # 使用缓动函数
                ease = progress * (2 - progress)
                current_x = int(start_x + (x - start_x) * ease)
                widget.place(x=current_x, y=y)
                widget.after(20, lambda: slide(step + 1))
            else:
                widget.place(x=x, y=y)

        slide(0)

    def _bind_mousewheel(self, widget):
        """为可滚动组件绑定鼠标滚轮事件"""
        def _on_mousewheel(event):
            # Windows系统使用 event.delta，Linux/Mac使用 event.num
            if event.delta:
                # Windows
                scroll_amount = -1 * (event.delta // 120)
            else:
                # Linux/Mac
                scroll_amount = -1 if event.num == 4 else 1

            # 滚动 Treeview
            widget.yview_scroll(scroll_amount, "units")

        def _on_mousewheel_b4(event):
            # Windows系统使用 event.delta
            scroll_amount = -1 * (event.delta // 120)
            widget.yview_scroll(scroll_amount, "units")

        # 绑定 Windows 系统的鼠标滚轮事件
        widget.bind("<MouseWheel>", _on_mousewheel, "+")
        # 绑定 Linux/Mac 系统的鼠标滚轮事件
        widget.bind("<Button-4>", lambda e: widget.yview_scroll(-1, "units"), "+")
        widget.bind("<Button-5>", lambda e: widget.yview_scroll(1, "units"), "+")

    def _bind_mousewheel_to_canvas(self, canvas):
        """为Canvas绑定鼠标滚轮事件"""
        def _on_mousewheel(event):
            # Windows系统使用 event.delta
            if event.delta:
                scroll_amount = -1 * (event.delta // 120)
            else:
                scroll_amount = -1 if event.num == 4 else 1
            canvas.yview_scroll(scroll_amount, "units")

        # 绑定 Windows 系统的鼠标滚轮事件
        canvas.bind("<MouseWheel>", _on_mousewheel, "+")
        # 绑定 Linux/Mac 系统的鼠标滚轮事件
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"), "+")
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"), "+")

    def _bind_mousewheel_to_listbox(self, listbox):
        """为Listbox绑定鼠标滚轮事件"""
        def _on_mousewheel(event):
            # Windows系统使用 event.delta
            if event.delta:
                scroll_amount = -1 * (event.delta // 120)
            else:
                scroll_amount = -1 if event.num == 4 else 1
            listbox.yview_scroll(scroll_amount, "units")

        # 绑定 Windows 系统的鼠标滚轮事件
        listbox.bind("<MouseWheel>", _on_mousewheel, "+")
        # 绑定 Linux/Mac 系统的鼠标滚轮事件
        listbox.bind("<Button-4>", lambda e: listbox.yview_scroll(-1, "units"), "+")
        listbox.bind("<Button-5>", lambda e: listbox.yview_scroll(1, "units"), "+")

    def _bind_mousewheel_to_text(self, text_widget):
        """为Text组件绑定鼠标滚轮事件"""
        def _on_mousewheel(event):
            # Windows系统使用 event.delta
            if event.delta:
                scroll_amount = -1 * (event.delta // 120)
            else:
                scroll_amount = -1 if event.num == 4 else 1
            text_widget.yview_scroll(scroll_amount, "units")

        # 绑定 Windows 系统的鼠标滚轮事件
        text_widget.bind("<MouseWheel>", _on_mousewheel, "+")
        # 绑定 Linux/Mac 系统的鼠标滚轮事件
        text_widget.bind("<Button-4>", lambda e: text_widget.yview_scroll(-1, "units"), "+")
        text_widget.bind("<Button-5>", lambda e: text_widget.yview_scroll(1, "units"), "+")

    def create_left_panel(self, parent):
        """创建左侧面板 - 日历和操作（宽松设计 - 删除学生名单）"""
        # 创建内容容器
        content_frame = tk.Frame(parent, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # 日历区域
        calendar_label = tk.Label(content_frame, text="📅 日历", 
                                font=('Microsoft YaHei', 13, 'bold'),
                                bg=self.colors['white'], fg=self.colors['fg'])
        calendar_label.pack(pady=(0, 12))
        
        calendar_frame = tk.Frame(content_frame, bg=self.colors['light_gray'])
        calendar_frame.pack(fill=tk.BOTH, expand=True)
        
        self.calendar = CalendarWidget(calendar_frame)
        # 设置日历颜色和回调
        self.calendar.colors = self.colors
        self.calendar.on_date_select = self.on_date_selected
        self.calendar.on_week_select = self.on_week_selected
        self.calendar.on_month_select = self.on_month_selected
        
        # 存储每个学生的请假类型选择
        self.student_leave_types = {}  # {name: "full" or "half" or None}
    
    def create_right_panel(self, parent):
        """创建右侧面板 - 学生列表和功能（宽松设计）"""
        # 创建内容容器
        content_frame = tk.Frame(parent, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        
        # 创建选项卡
        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # 绑定选项卡切换事件
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        # 录入选项卡
        input_tab = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(input_tab, text="📝 录入")
        self.create_input_tab(input_tab)

        # 统计选项卡
        stats_export_tab = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(stats_export_tab, text="📊 统计")
        self.create_stats_export_tab(stats_export_tab)

        # 设置选项卡
        settings_tab = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(settings_tab, text="⚙️ 设置")
        self.create_settings_tab(settings_tab)

        # 教程选项卡
        tutorial_tab = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(tutorial_tab, text="📖 教程")
        self.create_tutorial_tab(tutorial_tab)
    
    def create_input_tab(self, parent):
        """创建录入选项卡（宽松设计 - 添加全天半天选项）"""
        # 学生名单区域
        students_label = tk.Label(parent, text="👥 学生名单", 
                                font=('Microsoft YaHei', 12, 'bold'),
                                bg=self.colors['white'], fg=self.colors['fg'])
        students_label.pack(pady=(0, 10))
        
        # 学生列表（带全天/半天选项）
        students_frame = tk.Frame(parent, bg=self.colors['white'], relief='solid', borderwidth=1)
        students_frame.pack(fill=tk.X, pady=(0, 15))
        
        columns = ("name", "full", "half")
        self.students_tree = ttk.Treeview(students_frame, columns=columns, show="headings", height=14, selectmode="none")
        
        self.students_tree.heading("name", text="姓名")
        self.students_tree.heading("full", text="全天")
        self.students_tree.heading("half", text="半天")

        self.students_tree.column("name", width=150, anchor=tk.CENTER)
        self.students_tree.column("full", width=80, anchor=tk.CENTER)
        self.students_tree.column("half", width=80, anchor=tk.CENTER)
        
        # 添加表格线样式
        style = ttk.Style()
        style.configure("Treeview",
                       rowheight=25,
                       font=('Microsoft YaHei', 10),
                       background='white',
                       fieldbackground='white',
                       borderwidth=1,
                       relief='solid')
        style.configure("Treeview.Heading",
                       font=('Microsoft YaHei', 10, 'bold'),
                       background='#ECF0F1',
                       relief='solid',
                       borderwidth=1)
        style.map("Treeview",
                 background=[('selected', '#3498DB')],
                 foreground=[('selected', 'white')])

        # 为Treeview添加网格线
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])
        style.configure("Treeview", rowheight=25)
        style.map("Treeview",
                 background=[('selected', '#3498DB')],
                 foreground=[('selected', 'white')],
                 relief=[('active', 'groove')])

        # 为不同的行设置不同的背景色（每两行之间有灰色分隔线）
        style.configure("Treeview", rowheight=25)
        self.students_tree.configure(style="Treeview")
        self.students_tree.tag_configure('odd', background='white')
        self.students_tree.tag_configure('even', background='#E0E0E0')  # 深灰色
        
        students_scrollbar = ttk.Scrollbar(students_frame, orient=tk.VERTICAL, command=self.students_tree.yview)
        self.students_tree.config(yscrollcommand=students_scrollbar.set)

        self.students_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        students_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.students_tree.bind("<Button-1>", self.on_student_click)
        # 添加鼠标滚轮滚动
        self._bind_mousewheel(self.students_tree)
        
        # 常请假名单区域
        frequent_label = tk.Label(parent, text="⚠️ 常请假名单",
                                font=('Microsoft YaHei', 12, 'bold'),
                                bg=self.colors['white'], fg=self.colors['danger'])
        frequent_label.pack(pady=(0, 10))
        
        frequent_frame = tk.Frame(parent, bg=self.colors['white'], relief='solid', borderwidth=1)
        frequent_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.frequent_tree = ttk.Treeview(frequent_frame, columns=columns, show="headings", height=8, selectmode="none")
        
        self.frequent_tree.heading("name", text="姓名")
        self.frequent_tree.heading("full", text="全天")
        self.frequent_tree.heading("half", text="半天")

        self.frequent_tree.column("name", width=150, anchor=tk.CENTER)
        self.frequent_tree.column("full", width=80, anchor=tk.CENTER)
        self.frequent_tree.column("half", width=80, anchor=tk.CENTER)

        self.frequent_tree.configure(style="Treeview")
        self.frequent_tree.tag_configure('odd', background='white')
        self.frequent_tree.tag_configure('even', background='#E0E0E0')  # 深灰色

        frequent_scrollbar = ttk.Scrollbar(frequent_frame, orient=tk.VERTICAL, command=self.frequent_tree.yview)
        self.frequent_tree.config(yscrollcommand=frequent_scrollbar.set)

        self.frequent_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        frequent_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.frequent_tree.bind("<Button-1>", self.on_frequent_click)
        # 添加鼠标滚轮滚动
        self._bind_mousewheel(self.frequent_tree)
    
    def create_stats_export_tab(self, parent):
        """创建统计和导出选项卡（表格显示，自动生成）"""
        # 统计类型选择
        type_frame = tk.Frame(parent, bg=self.colors['white'])
        type_frame.pack(fill=tk.X, pady=(0, 12))
        
        type_label = tk.Label(type_frame, text="统计类型:", 
                           font=('Microsoft YaHei', 11, 'bold'),
                           bg=self.colors['white'], fg=self.colors['fg'])
        type_label.pack(side=tk.LEFT, padx=(0, 10))
        
        self.stats_type_var = tk.StringVar(value="current")
        current_radio = ttk.Radiobutton(type_frame, text="选择日期", variable=self.stats_type_var, value="current", command=self.on_stats_type_change)
        current_radio.pack(side=tk.LEFT, padx=(0, 10))

        week_radio = ttk.Radiobutton(type_frame, text="本周", variable=self.stats_type_var, value="week", command=self.on_stats_type_change)
        week_radio.pack(side=tk.LEFT, padx=(0, 10))

        month_radio = ttk.Radiobutton(type_frame, text="本月", variable=self.stats_type_var, value="month", command=self.on_stats_type_change)
        month_radio.pack(side=tk.LEFT, padx=(0, 10))
        
        custom_radio = ttk.Radiobutton(type_frame, text="自定义", variable=self.stats_type_var, value="custom", command=self.on_stats_type_change)
        custom_radio.pack(side=tk.LEFT)

        # 学生选择和日期范围选择（合并到一行）
        student_date_frame = tk.Frame(parent, bg=self.colors['white'])
        student_date_frame.pack(fill=tk.X, pady=(0, 12))

        # 学生选择
        student_label = tk.Label(student_date_frame, text="选择学生:",
                                font=('Microsoft YaHei', 10),
                                bg=self.colors['white'], fg=self.colors['fg'])
        student_label.pack(side=tk.LEFT, padx=(0, 8))

        self.selected_student_var = tk.StringVar(value="全部学生")
        self.student_combo = ttk.Combobox(student_date_frame, textvariable=self.selected_student_var,
                                         values=["全部学生"], state="readonly", width=20)
        self.student_combo.pack(side=tk.LEFT)
        self.student_combo.bind("<<ComboboxSelected>>", self.on_student_change)

        # 日期范围选择（自定义时显示）
        self.date_range_frame = tk.Frame(student_date_frame, bg=self.colors['white'])

        start_label = tk.Label(self.date_range_frame, text="开始日期:",
                              font=('Microsoft YaHei', 10),
                              bg=self.colors['white'], fg=self.colors['fg'])
        start_label.pack(side=tk.LEFT, padx=(20, 8))

        self.start_date_var = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        self.start_date_entry = tk.Entry(self.date_range_frame, textvariable=self.start_date_var, width=12,
                                   font=('Microsoft YaHei', 10),
                                   bg=self.colors['light_gray'], fg=self.colors['fg'],
                                   relief='solid', borderwidth=1)
        self.start_date_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.start_date_entry.bind('<Button-1>', self.on_date_entry_click)
        self.start_date_entry.bind('<KeyRelease>', self.on_date_entry_change)

        end_label = tk.Label(self.date_range_frame, text="结束日期:",
                            font=('Microsoft YaHei', 10),
                            bg=self.colors['white'], fg=self.colors['fg'])
        end_label.pack(side=tk.LEFT, padx=(0, 8))

        self.end_date_var = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        self.end_date_entry = tk.Entry(self.date_range_frame, textvariable=self.end_date_var, width=12,
                                 font=('Microsoft YaHei', 10),
                                 bg=self.colors['light_gray'], fg=self.colors['fg'],
                                 relief='solid', borderwidth=1)
        self.end_date_entry.pack(side=tk.LEFT)
        self.end_date_entry.bind('<Button-1>', self.on_date_entry_click)
        self.end_date_entry.bind('<KeyRelease>', self.on_date_entry_change)

        # 隐藏日期范围选择框
        self.date_range_frame.pack_forget()

        # 操作按钮
        button_frame = tk.Frame(parent, bg=self.colors['white'])
        button_frame.pack(fill=tk.X, pady=(0, 12))
        
        refresh_btn = tk.Button(button_frame, text="🔄 刷新", 
                             command=self.refresh_stats,
                             bg=self.colors['accent'], fg=self.colors['white'],
                             font=('Microsoft YaHei', 10, 'bold'), relief='flat',
                             padx=16, pady=8, cursor='hand2')
        refresh_btn.pack(side=tk.LEFT, padx=(0, 10))
        self._add_button_hover_effect(refresh_btn, self.colors['accent'], self.colors['accent_hover'])
        
        export_btn = tk.Button(button_frame, text="📥 导出Excel",
                              command=self.export_to_excel,
                              bg=self.colors['success'], fg=self.colors['white'],
                              font=('Segoe UI Symbol', 10, 'bold'), relief='flat',
                              padx=16, pady=8, cursor='hand2')
        export_btn.pack(side=tk.LEFT)
        self._add_button_hover_effect(export_btn, self.colors['success'], '#229954')
        
        # 统计结果显示（表格）
        stats_label = tk.Label(parent, text="📊 统计结果",
                              font=('Microsoft YaHei', 12, 'bold'),
                              bg=self.colors['white'], fg=self.colors['fg'])
        stats_label.pack(pady=(0, 10))

        # 使用Canvas绘制表格以支持动态行高
        stats_canvas_frame = tk.Frame(parent, bg=self.colors['white'])
        stats_canvas_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 创建Canvas和滚动条
        self.stats_canvas = tk.Canvas(stats_canvas_frame, bg='white', highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(stats_canvas_frame, orient=tk.VERTICAL, command=self.stats_canvas.yview)
        h_scrollbar = ttk.Scrollbar(stats_canvas_frame, orient=tk.HORIZONTAL, command=self.stats_canvas.xview)

        self.stats_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # 布局
        self.stats_canvas.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        stats_canvas_frame.grid_rowconfigure(0, weight=1)
        stats_canvas_frame.grid_columnconfigure(0, weight=1)

        # 绑定鼠标滚轮
        self._bind_mousewheel_to_canvas(self.stats_canvas)
        
        # 导出进度
        self.export_progress = ttk.Progressbar(parent, mode='determinate')
        self.export_progress.pack(fill=tk.X, pady=(8, 0))
        
        self.export_status_label = ttk.Label(parent, text="")
        self.export_status_label.pack(pady=(8, 0))
        
        # 自动生成统计
        self.refresh_stats()

        # 绑定窗口大小改变事件,刷新表格
        # 使用防抖优化，避免频繁触发
        self._last_window_width = self.root.winfo_width()
        self._last_window_height = self.root.winfo_height()
        self._is_resizing = False
        self._last_resize_time = 0  # 记录最后一次调整时间
        self.root.bind('<Configure>', self.on_window_resize)

    def create_settings_tab(self, parent):
        """创建设置选项卡 - 分组布局"""
        # 创建主容器
        main_frame = tk.Frame(parent, bg=self.colors['white'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        # 标题
        title_frame = tk.Frame(main_frame, bg=self.colors['accent'], padx=20, pady=15)
        title_frame.pack(fill=tk.X, pady=(0, 25))

        title_label = tk.Label(title_frame, text="⚙️ 系统设置",
                              font=('Microsoft YaHei', 18, 'bold'),
                              bg=self.colors['accent'], fg=self.colors['white'])
        title_label.pack()

        # 常规设置分组
        general_frame = tk.LabelFrame(main_frame, text="  常规设置  ",
                                       font=('Microsoft YaHei', 13, 'bold'),
                                       bg=self.colors['white'], fg=self.colors['fg'],
                                       padx=20, pady=20)
        general_frame.pack(fill=tk.X, pady=(0, 20))

        # 开机自启Web服务器
        self.auto_start_web_var = tk.BooleanVar(value=True)
        auto_start_web_frame = tk.Frame(general_frame, bg=self.colors['white'])
        auto_start_web_frame.pack(fill=tk.X, pady=(0, 10))

        auto_start_web_check = tk.Checkbutton(auto_start_web_frame, text="开机自启Web服务器",
                                             variable=self.auto_start_web_var,
                                             font=('Microsoft YaHei', 12),
                                             bg=self.colors['white'], fg=self.colors['fg'],
                                             activebackground=self.colors['white'],
                                             selectcolor=self.colors['light_gray'])
        auto_start_web_check.pack(side=tk.LEFT)

        auto_start_web_desc = tk.Label(auto_start_web_frame, text="  (开机时自动启动Web服务器,方便手机访问)",
                                      font=('Microsoft YaHei', 10), fg=self.colors['fg'],
                                      bg=self.colors['white'])
        auto_start_web_desc.pack(side=tk.LEFT)

        # 备份设置分组
        backup_frame = tk.LabelFrame(main_frame, text="  备份设置  ",
                                      font=('Microsoft YaHei', 13, 'bold'),
                                      bg=self.colors['white'], fg=self.colors['fg'],
                                      padx=20, pady=20)
        backup_frame.pack(fill=tk.X, pady=(0, 20))

        # 自动备份频率
        backup_freq_frame = tk.Frame(backup_frame, bg=self.colors['white'])
        backup_freq_frame.pack(fill=tk.X, pady=(0, 15))

        backup_freq_label = tk.Label(backup_freq_frame, text="自动备份频率(天):",
                                    font=('Microsoft YaHei', 12),
                                    bg=self.colors['white'], fg=self.colors['fg'])
        backup_freq_label.pack(side=tk.LEFT)

        self.backup_freq_var = tk.IntVar(value=1)
        backup_freq_spinbox = tk.Spinbox(backup_freq_frame, from_=1, to=7,
                                        textvariable=self.backup_freq_var,
                                        width=8,
                                        font=('Microsoft YaHei', 11))
        backup_freq_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        backup_freq_desc = tk.Label(backup_freq_frame, text="  (每N天自动备份一次数据)",
                                  font=('Microsoft YaHei', 10), fg=self.colors['fg'],
                                  bg=self.colors['white'])
        backup_freq_desc.pack(side=tk.LEFT, padx=(10, 0))

        # 保留备份文件数量
        backup_delete_frame = tk.Frame(backup_frame, bg=self.colors['white'])
        backup_delete_frame.pack(fill=tk.X, pady=(0, 15))

        backup_delete_label = tk.Label(backup_delete_frame, text="保留备份文件数量:",
                                      font=('Microsoft YaHei', 12),
                                      bg=self.colors['white'], fg=self.colors['fg'])
        backup_delete_label.pack(side=tk.LEFT)

        self.backup_delete_var = tk.IntVar(value=10)
        backup_delete_spinbox = tk.Spinbox(backup_delete_frame, from_=1, to=999,
                                          textvariable=self.backup_delete_var,
                                          width=8,
                                          font=('Microsoft YaHei', 11))
        backup_delete_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        backup_delete_desc = tk.Label(backup_delete_frame, text="  (自动删除旧备份,只保留最新的N个文件)",
                                    font=('Microsoft YaHei', 10), fg=self.colors['fg'],
                                    bg=self.colors['white'])
        backup_delete_desc.pack(side=tk.LEFT, padx=(10, 0))

        # 常请假名单设置分组
        frequent_frame = tk.LabelFrame(main_frame, text="  常请假名单设置  ",
                                       font=('Microsoft YaHei', 13, 'bold'),
                                       bg=self.colors['white'], fg=self.colors['fg'],
                                       padx=20, pady=20)
        frequent_frame.pack(fill=tk.X, pady=(0, 20))

        # 统计天数设置
        frequent_days_frame = tk.Frame(frequent_frame, bg=self.colors['white'])
        frequent_days_frame.pack(fill=tk.X, pady=(0, 15))

        frequent_days_label = tk.Label(frequent_days_frame, text="统计天数(天):",
                                      font=('Microsoft YaHei', 12),
                                      bg=self.colors['white'], fg=self.colors['fg'])
        frequent_days_label.pack(side=tk.LEFT)

        self.frequent_days_var = tk.IntVar(value=5)
        frequent_days_spinbox = tk.Spinbox(frequent_days_frame, from_=1, to=30,
                                          textvariable=self.frequent_days_var,
                                          width=8,
                                          font=('Microsoft YaHei', 11),
                                          command=self.on_frequent_days_change)
        frequent_days_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        frequent_days_desc = tk.Label(frequent_days_frame, text="  (统计最近N天内的请假记录)",
                                    font=('Microsoft YaHei', 10), fg=self.colors['fg'],
                                    bg=self.colors['white'])
        frequent_days_desc.pack(side=tk.LEFT, padx=(10, 0))

        # 请假次数阈值设置
        frequent_count_frame = tk.Frame(frequent_frame, bg=self.colors['white'])
        frequent_count_frame.pack(fill=tk.X, pady=(0, 15))

        frequent_count_label = tk.Label(frequent_count_frame, text="请假次数阈值:",
                                       font=('Microsoft YaHei', 12),
                                       bg=self.colors['white'], fg=self.colors['fg'])
        frequent_count_label.pack(side=tk.LEFT)

        self.frequent_count_var = tk.IntVar(value=3)
        frequent_count_spinbox = tk.Spinbox(frequent_count_frame, from_=1, to=99,
                                           textvariable=self.frequent_count_var,
                                           width=8,
                                           font=('Microsoft YaHei', 11),
                                           command=self.on_frequent_count_change)
        frequent_count_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        frequent_count_desc = tk.Label(frequent_count_frame, text="  (请假次数≥N次的学生将出现在名单中)",
                                     font=('Microsoft YaHei', 10), fg=self.colors['fg'],
                                     bg=self.colors['white'])
        frequent_count_desc.pack(side=tk.LEFT, padx=(10, 0))

        # 备份按钮组
        backup_buttons_frame = tk.Frame(backup_frame, bg=self.colors['white'])
        backup_buttons_frame.pack(fill=tk.X, pady=(10, 0))

        # 手动备份按钮
        create_backup_btn = tk.Button(backup_buttons_frame, text="💾 立即备份",
                                     command=self.create_backup,
                                     bg=self.colors['success'], fg=self.colors['white'],
                                     font=('Segoe UI Symbol', 11, 'bold'), relief='flat',
                                     padx=20, pady=10, cursor='hand2')
        create_backup_btn.pack(side=tk.LEFT, padx=(0, 15))
        self._add_button_hover_effect(create_backup_btn, self.colors['success'], '#229954')

        # 备份导入按钮
        import_backup_btn = tk.Button(backup_buttons_frame, text="📥 备份导入",
                                     command=self.import_backup,
                                     bg=self.colors['accent'], fg=self.colors['white'],
                                     font=('Segoe UI Symbol', 11, 'bold'), relief='flat',
                                     padx=20, pady=10, cursor='hand2')
        import_backup_btn.pack(side=tk.LEFT)
        self._add_button_hover_effect(import_backup_btn, self.colors['accent'], self.colors['accent_hover'])

    def create_backup(self, is_auto=False):
        """创建备份"""
        try:
            # 检查数据文件夹是否存在
            data_dir = 'data'
            if not os.path.exists(data_dir):
                if not is_auto:
                    messagebox.showwarning("警告", "数据文件夹不存在!\n请先运行程序并添加学生或录入请假记录,然后再创建备份。")
                return False

            # 检查是否有数据文件
            data_files = [f for f in os.listdir(data_dir) if f.endswith('.json') and f != 'settings.json']
            if not data_files:
                if not is_auto:
                    messagebox.showwarning("警告", "没有找到数据文件!\n请先添加学生或录入请假记录,然后再创建备份。")
                return False

            # 检查备份文件夹是否存在
            backup_dir = 'backup'
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)

            # 生成备份文件名
            from datetime import datetime
            if is_auto:
                backup_filename = f"自动备份-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.zip"
            else:
                backup_filename = f"手动备份-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.zip"
            backup_path = os.path.join(backup_dir, backup_filename)

            # 创建ZIP文件
            import zipfile
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 添加数据文件(排除settings.json)
                for file in data_files:
                    file_path = os.path.join(data_dir, file)
                    if os.path.isfile(file_path):
                        zipf.write(file_path, os.path.basename(file_path))

            # 备份成功后,自动删除旧备份
            self.auto_delete_old_backups()

            # 显示备份成功信息
            if is_auto:
                backup_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.update_status(f"自动备份成功: {backup_time}")
            else:
                backup_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.update_status(f"手动备份成功: {backup_time}")

            return True
        except Exception as e:
            if is_auto:
                self.update_status(f"自动备份失败: {str(e)}")
            else:
                messagebox.showerror("错误", f"创建备份失败: {str(e)}")
            return False

    def auto_delete_old_backups(self):
        """自动删除旧备份,保留最新的N个"""
        try:
            backup_dir = 'backup'
            if not os.path.exists(backup_dir):
                return

            # 获取所有备份文件
            backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.zip')]

            # 获取保留数量
            keep_count = getattr(self, 'backup_delete_var', None)
            if keep_count is None:
                keep_count = 10  # 默认保留10个
            else:
                keep_count = keep_count.get()

            if len(backup_files) > keep_count:
                # 按创建时间排序,保留最新的N个
                backup_files_with_time = []
                for file in backup_files:
                    file_path = os.path.join(backup_dir, file)
                    creation_time = os.path.getctime(file_path)
                    backup_files_with_time.append((file, creation_time))

                # 按创建时间降序排序(最新的在前)
                backup_files_with_time.sort(key=lambda x: x[1], reverse=True)

                # 删除超过保留数量的旧备份
                files_to_delete = backup_files_with_time[keep_count:]
                for file, _ in files_to_delete:
                    file_path = os.path.join(backup_dir, file)
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        # 删除失败不影响自动备份
                        pass
        except Exception as e:
            # 删除失败不影响自动备份
            pass

    def check_auto_backup(self):
        """检查是否需要自动备份"""
        try:
            backup_dir = 'backup'
            if not os.path.exists(backup_dir):
                # 没有备份文件夹,需要创建备份
                return True

            # 获取所有备份文件
            backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.zip')]

            if not backup_files:
                # 没有备份文件,需要创建备份
                return True

            # 获取最后一个备份文件的创建时间
            backup_files.sort(key=lambda x: os.path.getctime(os.path.join(backup_dir, x)))
            last_backup_file = backup_files[-1]
            last_backup_path = os.path.join(backup_dir, last_backup_file)
            last_backup_time = os.path.getctime(last_backup_path)

            # 获取自动备份频率
            backup_freq = getattr(self, 'backup_freq_var', None)
            if backup_freq is None:
                backup_freq = 1  # 默认1天
            else:
                backup_freq = backup_freq.get()

            # 计算距离上次备份的天数
            current_time = datetime.datetime.now().timestamp()
            days_since_last_backup = (current_time - last_backup_time) / (24 * 60 * 60)

            # 如果距离上次备份超过设定的天数,需要备份
            if days_since_last_backup >= backup_freq:
                return True
            else:
                return False
        except Exception as e:
            # 检查失败,不进行自动备份
            return False

    def import_backup(self):
        """导入备份"""
        # 检查备份文件夹是否存在
        backup_dir = 'backup'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # 获取备份文件列表
        backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.zip')]

        if not backup_files:
            messagebox.showinfo("提示", "没有找到备份文件!")
            return

        # 创建备份文件选择对话框
        dialog = tk.Toplevel(self.root)
        dialog.title("选择备份文件")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')

        # 标题
        tk.Label(dialog, text="选择要恢复的备份文件:",
                font=('Microsoft YaHei', 11, 'bold'),
                bg=self.colors['white'], fg=self.colors['fg']).pack(pady=10)

        # 备份文件列表
        listbox = tk.Listbox(dialog, height=10, width=40)
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=listbox.yview)
        listbox.config(yscrollcommand=scrollbar.set)

        # 按文件创建时间排序(最新的在前)
        backup_files_with_time = []
        for file in backup_files:
            file_path = os.path.join(backup_dir, file)
            creation_time = os.path.getctime(file_path)
            backup_files_with_time.append((file, creation_time))

        # 按创建时间降序排序(最新的在前)
        backup_files_with_time.sort(key=lambda x: x[1], reverse=True)

        for backup_file, _ in backup_files_with_time:
            listbox.insert(tk.END, backup_file)

        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(15, 0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 15), pady=10)

        # 添加鼠标滚轮滚动
        self._bind_mousewheel_to_listbox(listbox)

        def on_import():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个备份文件!")
                return

            selected_file = listbox.get(selection[0])
            backup_path = os.path.join(backup_dir, selected_file)

            # 确认对话框
            if messagebox.askyesno("警告", f"确定要恢复备份 '{selected_file}' 吗?\n当前数据将被覆盖!"):
                try:
                    # 解压备份文件到data文件夹
                    import zipfile
                    data_dir = 'data'
                    if not os.path.exists(data_dir):
                        os.makedirs(data_dir)

                    with zipfile.ZipFile(backup_path, 'r') as zip_ref:
                        zip_ref.extractall(data_dir)

                    messagebox.showinfo("成功", "备份已恢复!")
                    dialog.destroy()

                    # 重新加载数据
                    self.student_manager.load_students()
                    self.leave_manager.load_records()
                    self.refresh_students_list()
                    self.refresh_frequent_list()
                except Exception as e:
                    messagebox.showerror("错误", f"恢复备份失败: {str(e)}")

        def on_delete():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择要删除的备份文件!")
                return

            selected_file = listbox.get(selection[0])
            backup_path = os.path.join(backup_dir, selected_file)

            # 确认对话框
            if messagebox.askyesno("警告", f"确定要删除备份 '{selected_file}' 吗?\n此操作无法撤销!"):
                try:
                    os.remove(backup_path)
                    # 从列表中删除
                    listbox.delete(selection[0])
                    messagebox.showinfo("成功", "备份已删除!")
                except Exception as e:
                    messagebox.showerror("错误", f"删除备份失败: {str(e)}")

        # 按钮
        button_frame = tk.Frame(dialog, bg=self.colors['white'])
        button_frame.pack(pady=10)

        # 左侧按钮组(确定和删除上下排列)
        left_button_frame = tk.Frame(button_frame, bg=self.colors['white'])
        left_button_frame.pack(side=tk.LEFT, padx=5)

        tk.Button(left_button_frame, text="确定", command=on_import,
                bg=self.colors['success'], fg=self.colors['white'],
                font=('Microsoft YaHei', 10), relief='flat',
                padx=16, pady=6, cursor='hand2').pack(side=tk.TOP, pady=2)

        tk.Button(left_button_frame, text="删除", command=on_delete,
                bg=self.colors['danger'], fg=self.colors['white'],
                font=('Microsoft YaHei', 10), relief='flat',
                padx=16, pady=6, cursor='hand2').pack(side=tk.TOP, pady=2)

    def on_frequent_days_change(self):
        """统计天数改变时的处理"""
        try:
            days = self.frequent_days_var.get()
            count = self.frequent_count_var.get()
            if count > days:
                self.frequent_count_var.set(days)
            self.refresh_frequent_list()
        except:
            pass

    def on_frequent_count_change(self):
        """请假次数阈值改变时的处理"""
        try:
            days = self.frequent_days_var.get()
            count = self.frequent_count_var.get()
            if count > days:
                self.frequent_count_var.set(days)
            self.refresh_frequent_list()
        except:
            pass

    def save_settings(self):
        """保存设置到文件"""
        try:
            settings = {
                'auto_start_web': self.auto_start_web_var.get(),
                'backup_freq': self.backup_freq_var.get(),
                'backup_delete': self.backup_delete_var.get(),
                'frequent_days': self.frequent_days_var.get(),
                'frequent_count': self.frequent_count_var.get()
            }
            settings_file = os.path.join('data', 'settings.json')
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            pass

    def load_settings(self):
        """从文件加载设置"""
        try:
            settings_file = os.path.join('data', 'settings.json')
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    if 'auto_start_web' in settings:
                        self.auto_start_web_var.set(settings['auto_start_web'])
                    if 'backup_freq' in settings:
                        self.backup_freq_var.set(settings['backup_freq'])
                    if 'backup_delete' in settings:
                        self.backup_delete_var.set(settings['backup_delete'])
                    if 'frequent_days' in settings:
                        self.frequent_days_var.set(settings['frequent_days'])
                    if 'frequent_count' in settings:
                        self.frequent_count_var.set(settings['frequent_count'])
        except Exception as e:
            pass

    def create_tutorial_tab(self, parent):
        """创建教程选项卡 - 四格布局"""
        # 创建主容器
        main_frame = tk.Frame(parent, bg=self.colors['white'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_frame = tk.Frame(main_frame, bg=self.colors['accent'])
        title_frame.pack(fill=tk.X)

        title_label = tk.Label(title_frame, text="使用教程",
                               font=('Microsoft YaHei UI', 18, 'bold'),
                               bg=self.colors['accent'], fg=self.colors['white'])
        title_label.pack(pady=12)

        version_label = tk.Label(title_frame, text="版本：v1.0.1 | 作者：112114141",
                               font=('Microsoft YaHei UI', 10),
                               bg=self.colors['accent'], fg=self.colors['white'])
        version_label.pack(pady=(0, 12))

        # 创建四格容器
        panels_frame = tk.Frame(main_frame, bg=self.colors['white'])
        panels_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # 创建教程卡片
        tutorials = [
            {
                "icon": "👥",
                "title": "学生名单管理",
                "color": self.colors['accent'],
                "items": [
                    "添加单个学生：点击顶部「添加学生」按钮",
                    "批量导入学生：点击「导入学生」，每行一个姓名",
                    "删除学生：点击「删除学生」，可多选删除"
                ]
            },
            {
                "icon": "📝",
                "title": "请假录入",
                "color": self.colors['success'],
                "items": [
                    "选择日期：在左侧日历中点击选择日期",
                    "选择学生：点击「全天」或「半天」列进行选择",
                    "保存记录：点击顶部「保存」按钮",
                    "清空选择：点击顶部「清空」按钮"
                ]
            },
            {
                "icon": "📊",
                "title": "统计分析",
                "color": self.colors['warning'],
                "items": [
                    "选择统计类型：日期/本周/本月/自定义",
                    "选择学生：全部学生或单个学生",
                    "查看统计结果：自动生成表格显示",
                    "导出Excel：点击「导出Excel」按钮"
                ]
            },
            {
                "icon": "💡",
                "title": "注意事项",
                "color": self.colors['fg'],
                "items": [
                    "周一至周六为上学日，周日为休息日",
                    "数据自动保存，月初不清零",
                    "可编辑和修改历史数据",
                    "关闭程序时会提示保存未保存的数据"
                ]
            }
        ]

        # 创建四格布局
        for i, tutorial in enumerate(tutorials):
            row = i // 2
            col = i % 2

            # 卡片容器
            card = tk.Frame(panels_frame, bg=tutorial['color'], relief='raised', borderwidth=2)
            card.grid(row=row, column=col, sticky="nsew", padx=8, pady=8)

            # 配置网格权重
            panels_frame.grid_rowconfigure(row, weight=1)
            panels_frame.grid_columnconfigure(col, weight=1)

            # 卡片标题
            title_frame = tk.Frame(card, bg=tutorial['color'])
            title_frame.pack(fill=tk.X, padx=15, pady=(12, 8))

            title_label = tk.Label(title_frame, text=f"{tutorial['icon']} {tutorial['title']}",
                                  font=('Microsoft YaHei UI', 14, 'bold'),
                                  bg=tutorial['color'], fg=self.colors['white'])
            title_label.pack(anchor='w')

            # 卡片内容
            content_frame = tk.Frame(card, bg=self.colors['white'])
            content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))

            for item in tutorial['items']:
                item_label = tk.Label(content_frame, text=f"• {item}",
                                     font=('Microsoft YaHei UI', 11),
                                     bg=self.colors['white'], fg=self.colors['fg'],
                                     anchor='w', justify='left')
                item_label.pack(fill=tk.X, pady=5, padx=10)

    def _create_tutorial_panel(self, parent, row, col, title, items, color):
        """创建教程面板 - 修复emoji对齐问题"""
        # 面板容器
        panel = tk.Frame(parent, bg=self.colors['white'], relief='solid', borderwidth=1)
        panel.grid(row=row, column=col, sticky="nsew", padx=8, pady=8)

        # 配置网格权重
        parent.grid_rowconfigure(row, weight=1)
        parent.grid_columnconfigure(col, weight=1)

        # 面板标题
        title_bg = tk.Frame(panel, bg=color)
        title_bg.pack(fill=tk.X)

        title_label = tk.Label(title_bg, text=title,
                              font=('Microsoft YaHei UI', 13, 'bold'),
                              bg=color, fg=self.colors['white'])
        title_label.pack(pady=10)

        # 面板内容
        content_frame = tk.Frame(panel, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        for item in items:
            # 分离emoji和文字
            emoji_part = ""
            text_part = item

            # 检查是否有emoji在开头
            if item and len(item) > 0:
                # emoji通常是2-4个字符
                for i in range(min(4, len(item))):
                    if ord(item[i]) > 0x1F000:  # emoji的unicode范围
                        emoji_part = item[:i+1]
                        text_part = item[i+1:].lstrip()  # 移除emoji后的空格
                        break

            # 创建一行容器
            row_frame = tk.Frame(content_frame, bg=self.colors['white'])
            row_frame.pack(fill=tk.X, pady=5)

            # emoji标签
            if emoji_part:
                emoji_label = tk.Label(row_frame, text=emoji_part,
                                     font=('Segoe UI Emoji', 12),
                                     bg=self.colors['white'], fg=self.colors['fg'],
                                     width=2, anchor='w')
                emoji_label.pack(side=tk.LEFT)

            # 文字标签
            text_label = tk.Label(row_frame, text=text_part,
                                font=('Microsoft YaHei UI', 11),
                                bg=self.colors['white'], fg=self.colors['fg'],
                                anchor='w')
            text_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def load_initial_data(self):
        """加载初始数据"""
        # 加载设置
        self.root.after(100, self.load_settings)

        # 高亮日历日期（使用防抖优化）
        self._schedule_calendar_highlight()

        # 设置当前日期
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        self.date_var.set(today)
        self.calendar.set_selected_date(today)

        # 刷新学生列表（如果已创建）
        if hasattr(self, 'students_tree'):
            self.refresh_students_list()
            self.refresh_frequent_list()
            # 加载今天的请假记录
            self.load_leave_records(today)

        # 更新学生下拉框（如果需要）
        if hasattr(self, 'student_combo'):
            self.update_student_combos()

        # 延迟刷新统计界面，优化启动速度
        # 只有当用户切换到统计选项卡时才刷新
        if hasattr(self, 'stats_canvas'):
            self.root.after(300, self.refresh_stats)

        # 检查是否需要自动备份
        self.root.after(500, self.check_and_perform_auto_backup)

    def check_and_perform_auto_backup(self):
        """检查并执行自动备份"""
        try:
            if self.check_auto_backup():
                # 在后台线程中执行自动备份,避免阻塞UI
                threading.Thread(target=lambda: self.create_backup(is_auto=True), daemon=True).start()
        except Exception as e:
            pass
    
    def refresh_students_list(self):
        """刷新学生列表（显示全天半天选项）"""
        students = self.student_manager.get_students()

        # 清空表格
        for item in self.students_tree.get_children():
            self.students_tree.delete(item)

        # 添加学生到表格
        for i, student in enumerate(students):
            leave_type = self.student_leave_types.get(student, None)
            full_check = "✓" if leave_type == "full" else ""
            half_check = "✓" if leave_type == "half" else ""
            # 为每两行之间添加灰色分隔线
            tag = 'even' if (i + 1) % 2 == 0 else 'odd'
            self.students_tree.insert("", tk.END, values=(student, full_check, half_check), tags=(tag,))

    def refresh_frequent_list(self):
        """刷新常请假名单（显示全天半天选项）"""
        # 获取设置中的参数
        days = getattr(self, 'frequent_days_var', None)
        threshold = getattr(self, 'frequent_count_var', None)

        if days is None:
            days = 5
        else:
            days = days.get()

        if threshold is None:
            threshold = 3
        else:
            threshold = threshold.get()

        frequent_students = self.leave_manager.get_frequent_leavers(days=days, threshold=threshold)

        # 清空表格
        for item in self.frequent_tree.get_children():
            self.frequent_tree.delete(item)

        # 添加学生到表格
        for i, student in enumerate(frequent_students):
            leave_type = self.student_leave_types.get(student, None)
            full_check = "✓" if leave_type == "full" else ""
            half_check = "✓" if leave_type == "half" else ""
            # 为每两行之间添加灰色分隔线
            tag = 'even' if (i + 1) % 2 == 0 else 'odd'
            self.frequent_tree.insert("", tk.END, values=(student, full_check, half_check), tags=(tag,))

    def update_student_combos(self):
        """更新学生下拉框"""
        students = self.student_manager.get_students()
        # 如果table_student_combo存在，则更新它
        if hasattr(self, 'table_student_combo'):
            self.table_student_combo['values'] = students
    
    def on_student_click(self, event):
        """学生列表点击事件（直接点击切换全天/半天）"""
        # 获取点击的位置
        region = self.students_tree.identify("region", event.x, event.y)

        if region == "cell":
            # 获取点击的行和列
            item = self.students_tree.identify_row(event.y)
            column = self.students_tree.identify_column(event.x)

            if item:
                # 获取学生姓名
                values = self.students_tree.item(item, "values")
                student_name = values[0]

                # 获取列索引 (列名是 "#1", "#2", "#3" 之类的格式)
                col_index = int(column[1:]) - 1

                if col_index == 1:
                    # 点击全天列
                    current_type = self.student_leave_types.get(student_name, None)
                    if current_type == "full":
                        # 取消选择
                        del self.student_leave_types[student_name]
                    else:
                        # 选择全天
                        self.student_leave_types[student_name] = "full"
                    self.has_unsaved_changes = True  # 标记有未保存的修改
                    self.refresh_students_list()
                    self.refresh_frequent_list()
                elif col_index == 2:
                    # 点击半天列
                    current_type = self.student_leave_types.get(student_name, None)
                    if current_type == "half":
                        # 取消选择
                        del self.student_leave_types[student_name]
                    else:
                        # 选择半天
                        self.student_leave_types[student_name] = "half"
                    self.has_unsaved_changes = True  # 标记有未保存的修改
                    self.refresh_students_list()
                    self.refresh_frequent_list()
    
    def show_leave_type_dialog(self, student_name):
        """显示请假类型选择对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("选择请假类型")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        ttk.Label(dialog, text=f"学生: {student_name}").pack(pady=10)
        
        leave_type_var = tk.StringVar(value=self.student_leave_types.get(student_name, "none"))
        
        ttk.Radiobutton(dialog, text="全天", variable=leave_type_var, value="full").pack(anchor=tk.W, padx=20)
        ttk.Radiobutton(dialog, text="半天", variable=leave_type_var, value="half").pack(anchor=tk.W, padx=20)
        ttk.Radiobutton(dialog, text="不请假", variable=leave_type_var, value="none").pack(anchor=tk.W, padx=20)
        
        def on_confirm():
            leave_type = leave_type_var.get()
            if leave_type == "none":
                if student_name in self.student_leave_types:
                    del self.student_leave_types[student_name]
            else:
                self.student_leave_types[student_name] = leave_type
            
            self.refresh_students_list()
            self.refresh_frequent_list()
            dialog.destroy()
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="确定", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def on_frequent_click(self, event):
        """常请假名单点击事件（直接点击切换全天/半天）"""
        # 获取点击的位置
        region = self.frequent_tree.identify("region", event.x, event.y)

        if region == "cell":
            # 获取点击的行和列
            item = self.frequent_tree.identify_row(event.y)
            column = self.frequent_tree.identify_column(event.x)

            if item:
                # 获取学生姓名
                values = self.frequent_tree.item(item, "values")
                student_name = values[0]

                # 获取列索引 (列名是 "#1", "#2", "#3" 之类的格式)
                col_index = int(column[1:]) - 1

                if col_index == 1:
                    # 点击全天列
                    current_type = self.student_leave_types.get(student_name, None)
                    if current_type == "full":
                        # 取消选择
                        del self.student_leave_types[student_name]
                    else:
                        # 选择全天
                        self.student_leave_types[student_name] = "full"
                    self.has_unsaved_changes = True  # 标记有未保存的修改
                    self.refresh_students_list()
                    self.refresh_frequent_list()
                elif col_index == 2:
                    # 点击半天列
                    current_type = self.student_leave_types.get(student_name, None)
                    if current_type == "half":
                        # 取消选择
                        del self.student_leave_types[student_name]
                    else:
                        # 选择半天
                        self.student_leave_types[student_name] = "half"
                    self.has_unsaved_changes = True  # 标记有未保存的修改
                    self.refresh_students_list()
                    self.refresh_frequent_list()
    
    def _animate_selection_feedback(self):
        """选中反馈动画"""
        # 简单的闪烁效果
        pass
    
    def show_add_student_dialog(self):
        """显示添加学生对话框（居中显示）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加学生")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        ttk.Label(dialog, text="学生姓名:").pack(pady=10)
        
        name_var = tk.StringVar()
        name_entry = ttk.Entry(dialog, textvariable=name_var, width=20)
        name_entry.pack(pady=5)
        name_entry.focus()
        
        def add_student():
            name = name_var.get().strip()
            if name:
                if self.student_manager.add_student(name):
                    messagebox.showinfo("成功", f"已添加学生: {name}")
                    self.refresh_students_list()
                    self.update_student_combos()
                    dialog.destroy()
                else:
                    messagebox.showwarning("警告", "该学生已存在或姓名为空")
            else:
                messagebox.showwarning("警告", "请输入学生姓名")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="确定", command=add_student).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        name_entry.bind('<Return>', lambda e: add_student())
    
    def show_batch_import_dialog(self):
        """显示批量导入对话框（居中显示）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("批量导入学生")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        ttk.Label(dialog, text="每行输入一个学生姓名:").pack(pady=10)
        
        text_area = tk.Text(dialog, height=10, width=40)
        text_area.pack(pady=5, padx=10)
        text_area.focus()

        # 添加鼠标滚轮滚动
        self._bind_mousewheel_to_text(text_area)
        
        def import_students():
            content = text_area.get(1.0, tk.END).strip()
            names = [name.strip() for name in content.split('\n') if name.strip()]
            
            if names:
                count = self.student_manager.batch_import(names)
                messagebox.showinfo("成功", f"成功导入 {count} 个学生")
                self.refresh_students_list()
                self.update_student_combos()
                dialog.destroy()
            else:
                messagebox.showwarning("警告", "请输入学生姓名")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="导入", command=import_students).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def get_weekday(self, date_str: str) -> str:
        """获取星期几"""
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        return weekdays[date.weekday()]
    
    def on_date_selected(self, date_str: str):
        """日期选择事件"""
        # 检查是否有实际的未保存修改
        if self.has_unsaved_changes:
            if messagebox.askyesno("未保存的修改", "检测到有未保存的请假记录，是否保存？"):
                self.save_leave_record()

        self.date_var.set(date_str)
        self.load_leave_records(date_str)
        # 如果统计类型是"当前日期"，也刷新统计
        if hasattr(self, 'stats_type_var') and self.stats_type_var.get() == "current":
            self.refresh_stats()
    
    def on_week_selected(self, week_start: str, week_end: str):
        """周选择事件"""
        self.date_var.set(week_start)
        self.load_week_records(week_start, week_end)
    
    def on_month_selected(self, year: int, month: int):
        """月选择事件"""
        first_day = f"{year}-{month:02d}-01"
        self.date_var.set(first_day)
        self.load_month_records(year, month)
    
    def select_day_view(self):
        """选择日视图"""
        self.calendar.set_selected_date(self.date_var.get())
    
    def select_week_view(self):
        """选择周视图"""
        today = datetime.datetime.now()
        weekday = today.weekday()
        monday = today - datetime.timedelta(days=weekday)
        sunday = monday + datetime.timedelta(days=6)
        
        week_start = monday.strftime("%Y-%m-%d")
        week_end = sunday.strftime("%Y-%m-%d")

        self.calendar.select_week(week_start, week_end)

    def on_tab_changed(self, event):
        """选项卡切换事件"""
        # 检查是否有未保存的修改
        if self.has_unsaved_changes:
            if messagebox.askyesno("未保存的修改", "检测到有未保存的请假记录，是否保存？"):
                self.save_leave_record()

        # 如果切换到统计选项卡，刷新统计界面
        if hasattr(self, 'notebook'):
            current_tab = self.notebook.select()
            tab_text = self.notebook.tab(current_tab, "text")
            if "统计" in tab_text:
                # 延迟刷新，确保选项卡已经完全显示
                self.root.after(100, self.refresh_stats)

    def select_month_view(self):
        """选择月视图"""
        today = datetime.datetime.now()
        self.calendar.select_month(today.year, today.month)
    
    def load_week_records(self, week_start: str, week_end: str):
        """加载一周的请假记录"""
        # 清空学生列表选择
        self.student_leave_types.clear()
        if hasattr(self, 'students_tree'):
            self.refresh_students_list()
            self.refresh_frequent_list()

        # 更新统计信息
        self.update_week_stats(week_start, week_end)
    
    def load_month_records(self, year: int, month: int):
        """加载一个月的请假记录"""
        # 清空学生列表选择
        self.student_leave_types.clear()
        if hasattr(self, 'students_tree'):
            self.refresh_students_list()
            self.refresh_frequent_list()

        # 更新统计信息
        first_day = f"{year}-{month:02d}-01"
        if month == 12:
            last_day = f"{year + 1}-01-01"
        else:
            last_day = f"{year}-{month + 1:02d}-01"
        self.update_range_stats(first_day, last_day)
    
    def update_week_stats(self, week_start: str, week_end: str):
        """更新周统计（已废弃，统计信息现在在统计导出选项卡中）"""
        pass
    
    def update_range_stats(self, start_date: str, end_date: str):
        """更新范围统计（已废弃，统计信息现在在统计导出选项卡中）"""
        pass
    
    def load_leave_records(self, date_str: str):
        """加载某天的请假记录（显示全天/半天）"""
        records = self.leave_manager.get_leave_records(date_str)

        # 将请假记录加载到学生列表中
        self.student_leave_types.clear()
        for name, record in records.items():
            self.student_leave_types[name] = record["type"]

        # 重置未保存修改标志
        self.has_unsaved_changes = False

        # 刷新学生列表显示
        if hasattr(self, 'students_tree'):
            self.refresh_students_list()
            self.refresh_frequent_list()

    def show_remove_student_dialog(self):
        """显示删除学生对话框"""
        students = self.student_manager.get_students()
        if not students:
            messagebox.showinfo("提示", "没有学生可删除")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("删除学生")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="选择要删除的学生:").pack(pady=10)
        
        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, height=15)
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=listbox.yview)
        listbox.config(yscrollcommand=scrollbar.set)
        
        for student in students:
            listbox.insert(tk.END, student)
        
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))

        # 添加鼠标滚轮滚动
        self._bind_mousewheel_to_listbox(listbox)
        
        def remove_students():
            selected_indices = listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("警告", "请选择要删除的学生")
                return
            
            selected_students = [listbox.get(i) for i in selected_indices]
            
            if messagebox.askyesno("确认", f"确定要删除以下学生吗？\n{', '.join(selected_students)}"):
                for student in selected_students:
                    self.student_manager.remove_student(student)
                messagebox.showinfo("成功", f"已删除 {len(selected_students)} 个学生")
                self.refresh_students_list()
                self.update_student_combos()
                dialog.destroy()
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="删除", command=remove_students).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def select_date(self, date_type: str):
        """选择日期"""
        today = datetime.datetime.now()
        
        if date_type == "today":
            date_str = today.strftime("%Y-%m-%d")
        elif date_type == "week":
            # 本周一
            weekday = today.weekday()
            monday = today - datetime.timedelta(days=weekday)
            date_str = monday.strftime("%Y-%m-%d")
        elif date_type == "month":
            # 本月第一天
            date_str = today.replace(day=1).strftime("%Y-%m-%d")
        else:
            return
        
        self.date_var.set(date_str)
        self.calendar.set_selected_date(date_str)
        self.load_leave_records(date_str)
    
    def save_leave_record(self):
        """保存请假记录（改进版 - 添加超时保护和事务机制）"""
        # 防止重复点击 - 使用锁机制
        if self.is_saving:
            return

        self.is_saving = True

        # 立即禁用保存按钮，防止重复点击
        if hasattr(self, 'save_btn'):
            self.save_btn.config(state='disabled')

        # 添加超时保护机制（10秒后自动恢复）
        timeout_id = self.root.after(10000, self._reenable_save_button)

        try:
            date_str = self.date_var.get()

            if not date_str:
                messagebox.showwarning("警告", "请选择日期")
                self._reenable_save_button()
                self.root.after_cancel(timeout_id)  # 取消超时定时器
                return

            # 验证日期格式
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("警告", "日期格式无效")
                self._reenable_save_button()
                self.root.after_cancel(timeout_id)
                return

            # 获取所有选择了请假类型的学生
            selected_students = [(name, leave_type) for name, leave_type in self.student_leave_types.items()
                               if leave_type is not None]

            # 使用事务方式更新数据
            self._update_leave_records_with_transaction(date_str, selected_students)

            # 清空选择
            self.student_leave_types.clear()
            self.has_unsaved_changes = False

            # 刷新界面
            self.load_leave_records(date_str)

            # 高亮日历（使用防抖优化）
            self._schedule_calendar_highlight()

            # 添加成功动画（缩短动画时间到800毫秒）
            if selected_students:
                self._animate_success(f"已保存 {len(selected_students)} 个学生的请假记录")
            else:
                self._animate_success("已清空该日期的请假记录")

            # 取消超时定时器
            self.root.after_cancel(timeout_id)

        except Exception as e:
            # 记录错误日志
            print(f"保存错误: {str(e)}")
            messagebox.showerror("保存错误", f"保存失败：{str(e)}")
            self._reenable_save_button()
            self.root.after_cancel(timeout_id)
            return
        finally:
            # 确保按钮在动画结束后重新启用（动画持续800毫秒）
            if hasattr(self, 'save_btn'):
                self.root.after(900, self._reenable_save_button)
    
    def _update_leave_records_with_transaction(self, date_str: str, selected_students: list):
        """使用事务方式更新请假记录"""
        # 临时存储旧数据，以便回滚
        old_records = self.leave_manager.records.copy()

        try:
            # 删除该日期的现有记录（不立即保存）
            if date_str in self.leave_manager.records:
                del self.leave_manager.records[date_str]

            # 添加新记录（不立即保存）
            for student, leave_type in selected_students:
                if date_str not in self.leave_manager.records:
                    self.leave_manager.records[date_str] = {}
                self.leave_manager.records[date_str][student] = {"type": leave_type}

            # 统一保存到文件
            self.leave_manager.save_records()

        except Exception as e:
            # 回滚到旧数据
            self.leave_manager.records = old_records
            raise e

    def _reenable_save_button(self):
        """重新启用保存按钮"""
        self.is_saving = False
        if hasattr(self, 'save_btn'):
            self.save_btn.config(state='normal')
    
    def clear_selection(self):
        """清空选择"""
        self.student_leave_types.clear()
        self.refresh_students_list()
        self.refresh_frequent_list()
    
    def _format_text_with_linebreaks(self, text, max_chars_per_line=15):
        """将长文本格式化为多行显示，返回格式化后的文本和行数"""
        if not text or len(text) <= max_chars_per_line:
            return text, 1

        # 按逗号分割名字
        names = [name.strip() for name in text.split(",")]
        lines = []
        current_line = ""

        for name in names:
            if not current_line:
                current_line = name
            elif len(current_line + ", " + name) <= max_chars_per_line:
                current_line += ", " + name
            else:
                lines.append(current_line)
                current_line = name

        if current_line:
            lines.append(current_line)

        # 用换行符连接
        formatted_text = "\n".join(lines)
        return formatted_text, len(lines)

    def generate_statistics(self):
        """生成统计（使用Canvas绘制表格，支持动态行高）"""
        stats_type = self.stats_type_var.get()

        # 确定日期范围
        if stats_type == "current":
            start_date = self.date_var.get()
            end_date = start_date
        elif stats_type == "week":
            today = datetime.datetime.now()
            weekday = today.weekday()
            monday = today - datetime.timedelta(days=weekday)
            start_date = monday.strftime("%Y-%m-%d")
            end_date = (monday + datetime.timedelta(days=6)).strftime("%Y-%m-%d")
        elif stats_type == "month":
            today = datetime.datetime.now()
            start_date = today.replace(day=1).strftime("%Y-%m-%d")
            if today.month == 12:
                end_date = datetime.datetime(today.year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(today.year, today.month + 1, 1) - datetime.timedelta(days=1)
            end_date = end_date.strftime("%Y-%m-%d")
        else:  # custom
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()

        # 获取所有请假记录
        all_dates = self.leave_manager.get_all_dates()
        selected_student = self.selected_student_var.get()

        # 准备数据
        data = []
        if selected_student == "全部学生":
            # 按日期聚合统计
            date_stats = {}
            for date_str in all_dates:
                if start_date <= date_str <= end_date:
                    records = self.leave_manager.get_leave_records(date_str)
                    if date_str not in date_stats:
                        date_stats[date_str] = {"full": 0, "half": 0, "students": set(), "full_students": [], "half_students": []}
                    for name, record in records.items():
                        if record["type"] == "full":
                            date_stats[date_str]["full"] += 1
                            date_stats[date_str]["full_students"].append(name)
                        else:
                            date_stats[date_str]["half"] += 1
                            date_stats[date_str]["half_students"].append(name)
                        date_stats[date_str]["students"].add(name)

            for date_str in sorted(date_stats.keys()):
                weekday = self.get_weekday(date_str)
                count = len(date_stats[date_str]["students"])
                full_students = sorted(date_stats[date_str]["full_students"])
                half_students = sorted(date_stats[date_str]["half_students"])
                data.append({
                    "date": date_str,
                    "weekday": weekday,
                    "count": f"{count}人",
                    "full_students": full_students,
                    "half_students": half_students
                })
        else:
            # 单个学生统计
            for date_str in all_dates:
                if start_date <= date_str <= end_date:
                    records = self.leave_manager.get_leave_records(date_str)
                    if selected_student in records:
                        record = records[selected_student]
                        weekday = self.get_weekday(date_str)
                        full = record["type"] == "full"
                        half = record["type"] == "half"
                        data.append({
                            "date": date_str,
                            "weekday": weekday,
                            "count": selected_student,
                            "full_students": [selected_student] if full else [],
                            "half_students": [selected_student] if half else []
                        })

        # 使用Canvas绘制表格
        self._draw_stats_canvas(data)

    def _draw_stats_canvas(self, data):
        """使用Canvas绘制统计表格，支持动态行高，文字居中，宽度占满（性能优化版）"""
        # 保存当前数据，避免重复计算
        self._current_stats_data = data

        # 快速清空Canvas，不使用禁用/启用机制（实现实时效果）
        self.stats_canvas.delete("all")

        # 获取Canvas宽度
        canvas_width = self.stats_canvas.winfo_width()

        # 如果Canvas宽度太小，说明可能还没有正确渲染，延迟重绘
        if canvas_width < 50:
            # 保存数据以便延迟重绘
            self._pending_stats_data = data
            # 延迟30ms后重绘（更快响应）
            self.root.after(30, self._redraw_stats_canvas)
            return

        # 表格配置 - 根据Canvas宽度动态计算列宽
        col_widths = [
            canvas_width * 0.15,  # 日期
            canvas_width * 0.12,  # 星期
            canvas_width * 0.12,  # 人数/姓名
            canvas_width * 0.305, # 全天
            canvas_width * 0.305  # 半天
        ]
        row_height_base = 30
        line_height = 22
        x_padding = 10
        y_padding = 10

        # 判断是否为单个学生统计
        is_single_student = len(data) > 0 and "人" not in data[0]['count']

        # 根据统计类型设置第三个标题
        if is_single_student:
            headers = ["日期", "星期", "姓名", "全天", "半天"]
        else:
            headers = ["日期", "星期", "人数", "全天", "半天"]

        # 绘制表头
        y_pos = 0
        header_height = 40

        # 表头背景
        self.stats_canvas.create_rectangle(0, 0, canvas_width, header_height, fill='#4472C4', outline='')

        # 表头文字
        x_pos = 0
        for i, (header, width) in enumerate(zip(headers, col_widths)):
            self.stats_canvas.create_text(
                x_pos + width // 2,
                header_height // 2,
                text=header,
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )
            x_pos += width

        # 绘制表头边框
        x_pos = 0
        for width in col_widths:
            # 绘制垂直线
            self.stats_canvas.create_line(x_pos, 0, x_pos, header_height, fill='#FFFFFF', width=2)
            x_pos += width
        # 绘制底部水平线
        self.stats_canvas.create_line(0, header_height, canvas_width, header_height, fill='#FFFFFF', width=2)

        # 绘制数据行
        y_pos = header_height

        # 统计全天和半天的次数（用于单个学生统计）
        total_full_count = 0
        total_half_count = 0

        # 确保Canvas已完全更新后再绘制
        self.stats_canvas.update_idletasks()

        for row_data in data:
            weekday = row_data['weekday']

            # 确定背景色
            if weekday == "周六":
                bg_color = '#FFE699'
            elif weekday == "周日":
                bg_color = '#FFC7CE'
            else:
                bg_color = '#D9E1F2'

            # 计算需要的行数
            full_text = ", ".join(row_data['full_students'])
            half_text = ", ".join(row_data['half_students'])

            full_lines = self._count_lines(full_text, 20)
            half_lines = self._count_lines(half_text, 20)
            max_lines = max(full_lines, half_lines, 1)

            # 计算行高
            row_height = row_height_base + (max_lines - 1) * line_height

            # 绘制行背景
            self.stats_canvas.create_rectangle(0, y_pos, canvas_width, y_pos + row_height, fill=bg_color, outline='')

            # 绘制单元格内容
            x_pos = 0

            # 日期
            self.stats_canvas.create_text(
                x_pos + col_widths[0] // 2,
                y_pos + row_height // 2,
                text=row_data['date'],
                fill='#2C3E50',
                font=('Microsoft YaHei UI', 10)
            )
            x_pos += col_widths[0]

            # 星期
            self.stats_canvas.create_text(
                x_pos + col_widths[1] // 2,
                y_pos + row_height // 2,
                text=row_data['weekday'],
                fill='#2C3E50',
                font=('Microsoft YaHei UI', 10)
            )
            x_pos += col_widths[1]

            # 人数
            self.stats_canvas.create_text(
                x_pos + col_widths[2] // 2,
                y_pos + row_height // 2,
                text=row_data['count'],
                fill='#2C3E50',
                font=('Microsoft YaHei UI', 10)
            )
            x_pos += col_widths[2]

            # 全天（多行文本，居中）
            if is_single_student:
                # 单个学生统计，显示打钩
                if full_text:
                    self.stats_canvas.create_text(
                        x_pos + col_widths[3] // 2,
                        y_pos + row_height // 2,
                        text="✓",
                        fill='#2C3E50',
                        font=('Microsoft YaHei UI', 16, 'bold')
                    )
                    total_full_count += 1
            else:
                # 全部学生统计，显示学生名单
                if full_text:
                    # 计算多行文本的总高度
                    total_text_height = full_lines * line_height
                    # 计算起始Y坐标，使文本在单元格中完全居中
                    start_y = y_pos + (row_height - total_text_height) // 2
                    self._draw_multiline_text_centered(
                        self.stats_canvas,
                        full_text,
                        x_pos,
                        start_y,
                        col_widths[3],
                        line_height,
                        20
                    )

            x_pos += col_widths[3]

            # 半天（多行文本，居中）
            if is_single_student:
                # 单个学生统计，显示打钩
                if half_text:
                    self.stats_canvas.create_text(
                        x_pos + col_widths[4] // 2,
                        y_pos + row_height // 2,
                        text="✓",
                        fill='#2C3E50',
                        font=('Microsoft YaHei UI', 16, 'bold')
                    )
                    total_half_count += 1
            else:
                # 全部学生统计，显示学生名单
                if half_text:
                    # 计算多行文本的总高度
                    total_text_height = half_lines * line_height
                    # 计算起始Y坐标，使文本在单元格中完全居中
                    start_y = y_pos + (row_height - total_text_height) // 2
                    self._draw_multiline_text_centered(
                        self.stats_canvas,
                        half_text,
                        x_pos,
                        start_y,
                        col_widths[4],
                        line_height,
                        20
                    )

            # 绘制单元格边框（增强网格线）
            x_pos = 0
            for width in col_widths:
                # 绘制垂直线
                self.stats_canvas.create_line(x_pos, y_pos, x_pos, y_pos + row_height, fill='#95A5A6', width=2)
                x_pos += width
            # 绘制水平线（底部）
            self.stats_canvas.create_line(0, y_pos + row_height, canvas_width, y_pos + row_height, fill='#95A5A6', width=2)

            y_pos += row_height

        # 如果是单个学生统计，添加汇总行
        if is_single_student and data:
            # 绘制汇总行背景
            summary_height = 40
            summary_bg_color = '#4472C4'
            self.stats_canvas.create_rectangle(0, y_pos, canvas_width, y_pos + summary_height, fill=summary_bg_color, outline='')

            # 绘制汇总行内容
            x_pos = 0

            # 前两列合并显示"合计"
            self.stats_canvas.create_text(
                x_pos + col_widths[0] // 2,
                y_pos + summary_height // 2,
                text="合计",
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )
            self.stats_canvas.create_text(
                x_pos + col_widths[0] + col_widths[1] // 2,
                y_pos + summary_height // 2,
                text="",
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )
            x_pos += col_widths[0] + col_widths[1]

            # 人数列显示学生姓名
            student_name = data[0]['count'] if data else ""
            self.stats_canvas.create_text(
                x_pos + col_widths[2] // 2,
                y_pos + summary_height // 2,
                text=student_name,
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )
            x_pos += col_widths[2]

            # 全天列显示统计次数
            self.stats_canvas.create_text(
                x_pos + col_widths[3] // 2,
                y_pos + summary_height // 2,
                text=f"{total_full_count}次",
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )
            x_pos += col_widths[3]

            # 半天列显示统计次数
            self.stats_canvas.create_text(
                x_pos + col_widths[4] // 2,
                y_pos + summary_height // 2,
                text=f"{total_half_count}次",
                fill='white',
                font=('Microsoft YaHei UI', 11, 'bold')
            )

            # 绘制汇总行边框
            x_pos = 0
            for width in col_widths:
                # 绘制垂直线
                self.stats_canvas.create_line(x_pos, y_pos, x_pos, y_pos + summary_height, fill='#FFFFFF', width=2)
                x_pos += width
            # 绘制底部水平线
            self.stats_canvas.create_line(0, y_pos + summary_height, canvas_width, y_pos + summary_height, fill='#FFFFFF', width=2)

            y_pos += summary_height

        # 设置Canvas滚动区域
        self.stats_canvas.config(scrollregion=(0, 0, canvas_width, y_pos))

        # 强制立即更新，实现实时效果
        self.stats_canvas.update_idletasks()

    def _redraw_stats_canvas(self):
        """延迟重绘统计表格"""
        if hasattr(self, '_pending_stats_data') and self._pending_stats_data:
            self._draw_stats_canvas(self._pending_stats_data)
            self._pending_stats_data = None

    def on_window_resize(self, event):
        """窗口大小改变事件 - 实时刷新UI和表格（平滑效果）"""
        # 只处理root窗口的Configure事件，忽略子组件的事件
        if event.widget != self.root:
            return

        # 获取当前窗口尺寸
        current_width = self.root.winfo_width()
        current_height = self.root.winfo_height()

        # 检查窗口尺寸是否真的改变了（避免窗口移动时触发）
        # 使用较小的阈值（3像素），提高响应灵敏度
        width_changed = abs(current_width - self._last_window_width) > 3
        height_changed = abs(current_height - self._last_window_height) > 3

        # 更新最后记录的尺寸
        self._last_window_width = current_width
        self._last_window_height = current_height

        # 如果尺寸没有显著变化，直接返回（避免窗口移动时触发）
        if not width_changed and not height_changed:
            return

        # 记录当前时间
        import time
        current_time = time.time()
        self._last_resize_time = current_time

        # 取消之前的刷新任务
        if hasattr(self, '_resize_timer') and self._resize_timer:
            self.root.after_cancel(self._resize_timer)
        if hasattr(self, '_resize_timer2') and self._resize_timer2:
            self.root.after_cancel(self._resize_timer2)

        # 立即执行一次快速刷新（无延迟），确保最大化和恢复时立即响应
        self._resize_timer = self.root.after(0, lambda: self._refresh_ui_with_time_check(current_time))

        # 延迟执行第二次刷新（100ms），确保动画完成后最终状态正确
        self._resize_timer2 = self.root.after(100, lambda: self._refresh_ui_with_time_check(current_time))

    def _refresh_stats_optimized(self):
        """优化的统计刷新函数 - 只重绘表格，不重新生成数据"""
        # 如果有缓存的数据，直接重绘
        if hasattr(self, '_current_stats_data') and self._current_stats_data:
            self._draw_stats_canvas(self._current_stats_data)
        else:
            # 如果没有缓存数据，则执行完整刷新
            self.refresh_stats()

    def _refresh_stats_optimized_with_time_check(self, trigger_time):
        """优化的统计刷新函数 - 包含时间检查，确保只在最后一次调整后刷新"""
        import time

        # 只在当前是最后一次触发时才刷新
        if trigger_time == self._last_resize_time:
            # 如果有缓存的数据，直接重绘
            if hasattr(self, '_current_stats_data') and self._current_stats_data:
                self._draw_stats_canvas(self._current_stats_data)
            else:
                # 如果没有缓存数据，则执行完整刷新
                self.refresh_stats()

    def _refresh_ui_with_time_check(self, trigger_time):
        """实时刷新UI和表格 - 包含时间检查"""
        import time

        # 只在当前是最后一次触发时才刷新
        if trigger_time == self._last_resize_time:
            # 强制更新所有UI组件的布局
            self.root.update_idletasks()

            # 如果在统计选项卡，实时刷新表格
            if hasattr(self, 'notebook'):
                current_tab = self.notebook.select()
                if current_tab:
                    tab_text = self.notebook.tab(current_tab, "text")
                    if "统计" in tab_text:
                        # 如果有缓存的数据，直接重绘
                        if hasattr(self, '_current_stats_data') and self._current_stats_data:
                            # 确保Canvas已经正确渲染
                            if hasattr(self, 'stats_canvas'):
                                # 强制更新Canvas尺寸
                                self.stats_canvas.update()
                                # 重绘表格
                                self._draw_stats_canvas(self._current_stats_data)
                        else:
                            # 如果没有缓存数据，重新生成统计数据
                            self.refresh_stats()

    def _continuous_refresh_during_resize(self, trigger_time):
        """在窗口大小变化期间持续刷新"""
        import time

        # 检查是否还在调整大小（最近50ms内有新的触发）
        if time.time() - self._last_resize_time < 0.05:
            # 继续刷新
            if hasattr(self, '_current_stats_data') and self._current_stats_data:
                self.root.update_idletasks()
                if hasattr(self, 'notebook'):
                    current_tab = self.notebook.select()
                    if current_tab:
                        tab_text = self.notebook.tab(current_tab, "text")
                        if "统计" in tab_text:
                            if hasattr(self, 'stats_canvas'):
                                self.stats_canvas.update()
                                self._draw_stats_canvas(self._current_stats_data)
            # 50ms后再次检查
            self.root.after(50, lambda: self._continuous_refresh_during_resize(trigger_time))

    def _schedule_calendar_highlight(self):
        """延迟更新日历高亮（防抖优化）"""
        # 取消之前的定时器
        if self._calendar_update_timer:
            self.root.after_cancel(self._calendar_update_timer)

        # 延迟200ms后更新（减少延迟，加快启动）
        self._calendar_update_timer = self.root.after(200, self._do_calendar_highlight)

    def _do_calendar_highlight(self):
        """执行日历高亮更新"""
        dates = self.leave_manager.get_all_dates()
        self.calendar.highlight_dates(dates)
        self._calendar_update_timer = None

    def _count_lines(self, text, max_chars_per_line):
        """计算文本需要的行数"""
        if not text:
            return 1

        names = [name.strip() for name in text.split(",")]
        lines = []
        current_line = ""

        for name in names:
            if not current_line:
                current_line = name
            elif len(current_line + ", " + name) <= max_chars_per_line:
                current_line += ", " + name
            else:
                lines.append(current_line)
                current_line = name

        if current_line:
            lines.append(current_line)

        return len(lines)

    def _draw_multiline_text(self, canvas, text, x, y, max_width, line_height, max_chars_per_line):
        """绘制多行文本"""
        if not text:
            return

        # 分割文本为多行
        names = [name.strip() for name in text.split(",")]
        lines = []
        current_line = ""

        for name in names:
            if not current_line:
                current_line = name
            elif len(current_line + ", " + name) <= max_chars_per_line:
                current_line += ", " + name
            else:
                lines.append(current_line)
                current_line = name

        if current_line:
            lines.append(current_line)

        # 绘制每一行
        for i, line in enumerate(lines):
            canvas.create_text(
                x,
                y + i * line_height + line_height // 2,
                text=line,
                fill='#2C3E50',
                font=('Microsoft YaHei', 9),
                anchor='w'
            )

    def _draw_multiline_text_centered(self, canvas, text, x, y, width, line_height, max_chars_per_line):
        """绘制居中的多行文本"""
        if not text:
            return

        # 分割文本为多行
        names = [name.strip() for name in text.split(",")]
        lines = []
        current_line = ""

        for name in names:
            if not current_line:
                current_line = name
            elif len(current_line + ", " + name) <= max_chars_per_line:
                current_line += ", " + name
            else:
                lines.append(current_line)
                current_line = name

        if current_line:
            lines.append(current_line)

        # 绘制每一行（居中）
        for i, line in enumerate(lines):
            canvas.create_text(
                x + width // 2,
                y + i * line_height + line_height // 2,
                text=line,
                fill='#2C3E50',
                font=('Microsoft YaHei UI', 9),
                anchor='center'
            )
    
    def on_stats_type_change(self, event=None):
        """统计类型改变事件"""
        stats_type = self.stats_type_var.get()
        
        # 显示或隐藏日期范围选择框
        if stats_type == "custom":
            self.date_range_frame.pack(fill=tk.X, pady=(0, 12))
        else:
            self.date_range_frame.pack_forget()
        
        # 刷新统计
        self.refresh_stats()
    
    def on_student_change(self, event=None):
        """学生选择改变事件"""
        # 刷新统计
        self.refresh_stats()
    
    def on_date_entry_click(self, event):
        """日期输入框点击事件 - 弹出日历选择"""
        # 确定是哪个输入框被点击了
        widget = event.widget
        if widget == self.start_date_entry:
            target_var = self.start_date_var
        elif widget == self.end_date_entry:
            target_var = self.end_date_var
        else:
            return

        # 弹出日期选择对话框
        self.show_date_picker_dialog(target_var)

    def on_date_entry_change(self, event):
        """日期输入框内容改变事件 - 自动刷新统计"""
        # 只有在自定义统计类型下才刷新
        if hasattr(self, 'stats_type_var') and self.stats_type_var.get() == "custom":
            self.refresh_stats()
    
    def show_date_picker_dialog(self, target_var):
        """显示日期选择对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("选择日期")
        dialog.geometry("300x320")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        # 创建日历
        calendar_frame = tk.Frame(dialog, bg=self.colors['white'])
        calendar_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 月份导航
        nav_frame = tk.Frame(calendar_frame, bg=self.colors['white'])
        nav_frame.pack(fill=tk.X, pady=(0, 10))
        
        def prev_month():
            current_month = calendar_combo.get()
            year, month = map(int, current_month.split('-'))
            month -= 1
            if month < 1:
                month = 12
                year -= 1
            calendar_combo.set(f"{year}-{month:02d}")
            update_calendar()
        
        def next_month():
            current_month = calendar_combo.get()
            year, month = map(int, current_month.split('-'))
            month += 1
            if month > 12:
                month = 1
                year += 1
            calendar_combo.set(f"{year}-{month:02d}")
            update_calendar()
        
        prev_btn = tk.Button(nav_frame, text="<", width=3, command=prev_month)
        prev_btn.pack(side=tk.LEFT)
        
        # 月份选择下拉框
        current_date = datetime.datetime.now()
        months = []
        for m in range(1, 13):
            months.append(f"{current_date.year}-{m:02d}")
        
        calendar_combo = ttk.Combobox(nav_frame, values=months, state="readonly", width=10)
        calendar_combo.set(f"{current_date.year}-{current_date.month:02d}")
        calendar_combo.pack(side=tk.LEFT, padx=5)
        calendar_combo.bind("<<ComboboxSelected>>", lambda e: update_calendar())
        
        next_btn = tk.Button(nav_frame, text=">", width=3, command=next_month)
        next_btn.pack(side=tk.LEFT)
        
        # 星期标题
        week_frame = tk.Frame(calendar_frame, bg=self.colors['white'])
        week_frame.pack(fill=tk.X, pady=(0, 5))
        
        weekdays = ["日", "一", "二", "三", "四", "五", "六"]
        for i, day in enumerate(weekdays):
            label = tk.Label(week_frame, text=day, width=4, font=("Arial", 9),
                           bg=self.colors['white'], fg=self.colors['fg'])
            label.grid(row=0, column=i, padx=1, pady=1)
        
        # 日历主体
        days_frame = tk.Frame(calendar_frame, bg=self.colors['white'])
        days_frame.pack(fill=tk.BOTH, expand=True)
        
        def update_calendar():
            # 清空现有按钮
            for widget in days_frame.winfo_children():
                widget.destroy()
            
            # 获取当前选择的年月
            current_month = calendar_combo.get()
            year, month = map(int, current_month.split('-'))
            
            # 获取该月第一天是星期几
            first_day = datetime.datetime(year, month, 1)
            start_weekday = first_day.weekday() + 1  # 0=周一, 6=周日, 转换为0=周日, 6=周六
            
            # 获取该月总天数
            if month == 12:
                next_month = datetime.datetime(year + 1, 1, 1)
            else:
                next_month = datetime.datetime(year, month + 1, 1)
            total_days = (next_month - first_day).days
            
            # 创建日历按钮
            day = 1
            for row in range(6):
                for col in range(7):
                    if row == 0 and col < start_weekday:
                        continue
                    if day > total_days:
                        break
                    
                    def select_date(d=day):
                        date_str = f"{year}-{month:02d}-{d:02d}"
                        target_var.set(date_str)
                        dialog.destroy()
                        # 选择日期后自动刷新统计
                        if hasattr(self, 'stats_type_var') and self.stats_type_var.get() == "custom":
                            self.refresh_stats()
                    
                    btn = tk.Button(days_frame, text=str(day), width=4, height=1,
                                   bg=self.colors['light_gray'], fg=self.colors['fg'],
                                   command=select_date)
                    btn.grid(row=row, column=col, padx=1, pady=1)
                    
                    day += 1

        update_calendar()

        # 按钮区域（已删除关闭按钮，选择日期后自动关闭）
        # button_frame = tk.Frame(dialog, bg=self.colors['white'])
        # button_frame.pack(fill=tk.X, padx=10, pady=10)
        #
        # close_btn = tk.Button(button_frame, text="关闭", command=dialog.destroy,
        #                     bg=self.colors['light_gray'], fg=self.colors['fg'],
        #                     font=('Microsoft YaHei', 9), relief='flat',
        #                     padx=15, pady=5, cursor='hand2')
        # close_btn.pack(side=tk.RIGHT)

    def refresh_stats(self):
        """刷新统计"""
        # 更新学生列表
        students = self.student_manager.get_students()
        current_selection = self.selected_student_var.get()
        self.student_combo['values'] = ["全部学生"] + students

        # 如果当前选择的学生不在列表中，则重置为"全部学生"
        if current_selection not in self.student_combo['values']:
            self.selected_student_var.set("全部学生")

        # 强制更新Canvas宽度
        if hasattr(self, 'stats_canvas'):
            self.stats_canvas.update()

        # 生成统计
        self.generate_statistics()
    
    def update_student_combos(self):
        """更新学生下拉框"""
        students = self.student_manager.get_students()
        if hasattr(self, 'student_combo'):
            self.student_combo['values'] = ["全部学生"] + students
    
    def export_to_excel(self):
        """导出到Excel（功能全面优化版 - 表格数据）"""
        # 收集表格中的数据
        table_data = []
        # 从Canvas重新生成数据
        stats_type = self.stats_type_var.get()

        # 确定日期范围
        if stats_type == "current":
            start_date = self.date_var.get()
            end_date = start_date
        elif stats_type == "week":
            today = datetime.datetime.now()
            weekday = today.weekday()
            monday = today - datetime.timedelta(days=weekday)
            start_date = monday.strftime("%Y-%m-%d")
            end_date = (monday + datetime.timedelta(days=6)).strftime("%Y-%m-%d")
        elif stats_type == "month":
            today = datetime.datetime.now()
            start_date = today.replace(day=1).strftime("%Y-%m-%d")
            if today.month == 12:
                end_date = datetime.datetime(today.year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.datetime(today.year, today.month + 1, 1) - datetime.timedelta(days=1)
            end_date = end_date.strftime("%Y-%m-%d")
        else:  # custom
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()

        # 获取所有请假记录
        all_dates = self.leave_manager.get_all_dates()
        selected_student = self.selected_student_var.get()

        # 准备数据
        if selected_student == "全部学生":
            # 按日期聚合统计
            date_stats = {}
            for date_str in all_dates:
                if start_date <= date_str <= end_date:
                    records = self.leave_manager.get_leave_records(date_str)
                    if date_str not in date_stats:
                        date_stats[date_str] = {"full": 0, "half": 0, "students": set(), "full_students": [], "half_students": []}
                    for name, record in records.items():
                        if record["type"] == "full":
                            date_stats[date_str]["full"] += 1
                            date_stats[date_str]["full_students"].append(name)
                        else:
                            date_stats[date_str]["half"] += 1
                            date_stats[date_str]["half_students"].append(name)
                        date_stats[date_str]["students"].add(name)

            for date_str in sorted(date_stats.keys()):
                weekday = self.get_weekday(date_str)
                count = len(date_stats[date_str]["students"])
                full_students = sorted(date_stats[date_str]["full_students"])
                half_students = sorted(date_stats[date_str]["half_students"])
                table_data.append({
                    "date": date_str,
                    "weekday": weekday,
                    "col3": f"{count}人",
                    "col4": ", ".join(full_students),
                    "col5": ", ".join(half_students)
                })
        else:
            # 单个学生统计
            for date_str in all_dates:
                if start_date <= date_str <= end_date:
                    records = self.leave_manager.get_leave_records(date_str)
                    if selected_student in records:
                        record = records[selected_student]
                        weekday = self.get_weekday(date_str)
                        full = record["type"] == "full"
                        half = record["type"] == "half"
                        table_data.append({
                            "date": date_str,
                            "weekday": weekday,
                            "col3": selected_student,
                            "col4": "✓" if full else "",
                            "col5": "✓" if half else ""
                        })

        if not table_data:
            messagebox.showwarning("警告", "没有数据可导出")
            return

        # 选择保存位置
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            title="选择保存位置"
        )

        if not file_path:
            return

        # 在新线程中导出
        self.export_progress['value'] = 0
        self.export_status_label.config(text="正在导出...")

        thread = threading.Thread(target=self._export_excel_thread, args=(file_path, table_data))
        thread.start()
    
    def _export_excel_thread(self, file_path: str, table_data):
        """Excel导出线程"""
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "请假记录"

            # 检查是全部学生统计还是单个学生统计
            is_all_students = self.selected_student_var.get() == "全部学生"

            # 根据统计类型设置表头
            if is_all_students:
                headers = ["日期", "星期", "人数", "全天", "半天"]
            else:
                headers = ["日期", "星期", "姓名", "全天", "半天"]
            ws.append(headers)

            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 设置列宽（增加第三列宽度以容纳多人名单）
            ws.column_dimensions[get_column_letter(1)].width = 15
            ws.column_dimensions[get_column_letter(2)].width = 10
            ws.column_dimensions[get_column_letter(3)].width = 50 if is_all_students else 15  # 人数多时增加宽度
            ws.column_dimensions[get_column_letter(4)].width = 50  # 全天名单也可能很长
            ws.column_dimensions[get_column_letter(5)].width = 50  # 半天名单也可能很长

            # 添加数据并设置样式
            weekday_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            saturday_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            sunday_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # 定义细边框样式（用于有数据的单元格）
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for row_num, data in enumerate(table_data, 2):
                # 添加数据
                ws.append([data["date"], data["weekday"], data["col3"], data["col4"], data["col5"]])

                # 设置颜色
                fill = None
                if data["weekday"] == "周六":
                    fill = saturday_fill
                elif data["weekday"] == "周日":
                    fill = sunday_fill
                else:
                    fill = weekday_fill

                # 判断该行是否有数据
                has_data = False
                if data["col3"] and str(data["col3"]).strip():
                    has_data = True
                if data["col4"] and str(data["col4"]).strip():
                    has_data = True
                if data["col5"] and str(data["col5"]).strip():
                    has_data = True

                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    # 所有列都使用居中对齐和自动换行
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    # 为有数据的单元格添加细边框
                    if has_data:
                        cell.border = thin_border

                # 更新进度
                progress = (row_num - 1) / len(table_data) * 100
                self.export_progress['value'] = progress

            # 如果是单个学生统计，添加合计行
            if not is_all_students and table_data:
                # 统计全天和半天次数
                total_full_count = 0
                total_half_count = 0
                for data in table_data:
                    if data["col4"] and str(data["col4"]).strip():
                        total_full_count += 1
                    if data["col5"] and str(data["col5"]).strip():
                        total_half_count += 1

                # 添加合计行
                summary_row = ws.max_row + 1
                ws.append(["", "", "", "", ""])

                # 设置合计行样式
                summary_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                summary_font = Font(bold=True, color="FFFFFF", size=11)

                # 第一列：合计
                cell = ws.cell(row=summary_row, column=1)
                cell.value = "合计"
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 第二列：空
                cell = ws.cell(row=summary_row, column=2)
                cell.fill = summary_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 第三列：学生姓名
                student_name = table_data[0]["col3"] if table_data else ""
                cell = ws.cell(row=summary_row, column=3)
                cell.value = student_name
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 第四列：全天次数
                cell = ws.cell(row=summary_row, column=4)
                cell.value = f"{total_full_count}次"
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 第五列：半天次数
                cell = ws.cell(row=summary_row, column=5)
                cell.value = f"{total_half_count}次"
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 调整行高以适应内容
            for row_num in range(2, ws.max_row + 1):
                max_lines = 1
                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        # 计算需要的行数
                        text = str(cell.value)
                        # 根据列宽估算每行能显示的字符数
                        if col_num == 3 or col_num == 4 or col_num == 5:
                            # 第3、4、5列列宽较大，每行约显示20个字符
                            chars_per_line = 20
                        else:
                            # 其他列列宽较小，每行约显示10个字符
                            chars_per_line = 10

                        # 计算需要的行数
                        lines = (len(text) + chars_per_line - 1) // chars_per_line
                        max_lines = max(max_lines, lines)

                # 根据最大行数设置行高（每行高度为15）
                if max_lines > 1:
                    ws.row_dimensions[row_num].height = 15 * max_lines

            # 保存文件
            wb.save(file_path)

            # 更新状态
            self.export_status_label.config(text="导出完成！")
            self.export_progress['value'] = 100

            # 显示成功动画
            self._animate_success(f"成功导出 {len(table_data)} 条记录")

        except Exception as e:
            self.export_status_label.config(text=f"导出失败: {str(e)}")
            messagebox.showerror("错误", f"导出失败: {str(e)}")


def main():
    """主函数"""
    root = tk.Tk()
    app = LeaveRecordApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
